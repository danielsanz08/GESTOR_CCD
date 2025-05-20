from django.shortcuts import render, redirect
from django.contrib.auth import update_session_auth_hash
from django.contrib import messages
from django.core.mail import send_mail
from django.conf import settings
from django.contrib.auth import authenticate, login
from libreria.forms import CustomUserForm, CustomUserEditForm, CustomPasswordChangeForm
from django.shortcuts import get_object_or_404
from .models import CustomUser
from django.contrib.auth import get_user_model
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.core.mail import send_mail
from django.shortcuts import render, redirect
from django.urls import reverse
from django.template.loader import render_to_string
from django.contrib.auth.forms import SetPasswordForm
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.db.models import Q
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import get_user_model
# Create your views here.
# Página de inicio
User = get_user_model()

def inicio(request):
    return render(request, 'index/index.html')


def crear_usuario(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Crear usuario', 'url': '/crear_usuario'},
    ]
    admin_exists = CustomUser.objects.exists()

    if request.method == 'POST':
        form = CustomUserForm(request.POST)

        if form.is_valid():
            try:
                user = form.save(commit=False)
                module = user.module
                role = user.role

                # Verificar el número de administradores en cada módulo
                if role == 'Administrador':
                    admin_count = CustomUser.objects.filter(
                        role='Administrador', module=module, is_active=True
                    ).count()

                    limits = {'Papeleria': 3, 'Cafeteria': 2, 'Centro de eventos': 1}
                    if module in limits and admin_count >= limits[module]:
                        messages.error(
                            request,
                            f"Ya existen {limits[module]} administradores en el módulo {module}."
                        )
                        return redirect('crear_usuario')

                # Guardar el usuario
                user.save()

                # Enviar correo a los administradores del mismo módulo
                admin_emails = CustomUser.objects.filter(
                    role='Administrador', module=module, is_active=True
                ).values_list('email', flat=True)

                cargo = request.POST.get("cargo", "").strip()
                email = user.email  # Se obtiene directamente del objeto user

                if admin_emails:
                    subject = f"Nuevo usuario creado en {module}"
                    message = (
                        f"Hola querido usuario,\n\n"
                        f"Por parte de Gestor CCD, te informamos que se ha creado un nuevo usuario.\n\n"
                        f"Información:\n\n"
                        f"Nombre: {user.username}\n"
                        f"Rol: {role}\n"
                        f"Módulo: {module}\n"
                        f"Cargo: {cargo}\n"
                        f"Email: {email}\n\n"
                        f"En caso de ser infiltrado, por favor te invitamos a desactivarlo.\n\n"
                        f"Muchas gracias por su atención.\n"
                        f"El director de Gestor CCD te desea un feliz día."
                    )
                    try:
                        send_mail(
                            subject,
                            message,
                            settings.DEFAULT_FROM_EMAIL,
                            list(admin_emails),
                            fail_silently=False,
                        )
                        print("Correo enviado correctamente.")
                    except Exception as e:
                        messages.error(request, f"No se pudo enviar el correo: {e}")

                messages.success(request, f"Usuario '{user.username}' creado exitosamente.")
                return redirect('libreria:inicio')

            except Exception as e:
                messages.error(request, f"Hubo un error al crear el usuario: {e}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = CustomUserForm()

    return render(request, 'usuario/crear_usuario.html', {
        'form': form,
        'admin_exists': admin_exists,
        'breadcrumbs': breadcrumbs
    })

def ver_usuario(request, user_id):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Ver usuario', 'url': reverse('libreria:ver_usuario', kwargs={'user_id': user_id})},
    ]
    usuario = get_object_or_404(CustomUser, id=user_id)
    return render(request, 'usuario/ver_perfil.html', {'usuario': usuario, 'breadcrumbs': breadcrumbs})


def editar_usuario(request, user_id):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Ver usuario', 'url': reverse('libreria:ver_usuario', kwargs={'user_id': user_id})},
        {'name': 'Editar usuario', 'url': reverse('libreria:editar_usuario', kwargs={'user_id': user_id})},
    ]
    usuario = get_object_or_404(CustomUser, id=user_id)

    if request.method == 'POST':
        form = CustomUserEditForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            return redirect('libreria:lista_usuarios')

    else:
        form = CustomUserEditForm(instance=usuario)

    return render(request, 'usuario/editar_usuario.html', {
        'form': form,
        'usuario': usuario,
        'breadcrumbs': breadcrumbs
    })


def lista_usuarios(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Listado de usuarios', 'url': '/lista_usuarios'},
    ]
    query = request.GET.get('q', '').strip()
    usuarios_lista = CustomUser.objects.filter(module='papeleria')  # Ordenar por ID

    if query:
        usuarios_lista = usuarios_lista.filter(
            Q(username__icontains=query) |
            Q(email__icontains=query) |
            Q(role__icontains=query) |
            Q(cargo__icontains=query) |
            Q(is_active__icontains=query)
        )

    paginator = Paginator(usuarios_lista, 5)  # Muestra 10 usuarios por página
    page_number = request.GET.get('page')
    usuarios = paginator.get_page(page_number)

    return render(request, 'usuario/lista_usuarios.html', {'usuarios': usuarios, 'breadcrumbs': breadcrumbs})


def cambiar_estado_usuario(request, user_id):
    if request.method == 'POST':
        user = get_object_or_404(CustomUser, id=user_id)
        user.is_active = 'is_active' in request.POST
        user.save()
        return redirect(reverse("libreria:lista_usuarios"))  # Redirige a la página de login

def cambiar_contraseña(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Cambiar Contraseña', 'url': reverse('libreria:cambiar_contraseña')},
    ]
    if request.method == 'POST':
        form = CustomPasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            return redirect('libreria:inicio')  # Redirige después de cambiar la contraseña
    else:
        form = CustomPasswordChangeForm(user=request.user)

    return render(request, 'usuario/cambiar_contraseña.html', {'form': form, 'breadcrumbs': breadcrumbs})
def password_reset_request(request):
    if request.method == "POST":
        email = request.POST.get("email")
        try:
            user = User.objects.get(email=email)
            token = default_token_generator.make_token(user)
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            reset_link = request.build_absolute_uri(
                reverse("libreria:password_reset_confirm", kwargs={"uidb64": uid, "token": token})
            )

            subject = "Restablecer tu contraseña"
            message = render_to_string("password_reset_email.html", {"reset_link": reset_link})

            # Aquí NO capturamos errores de red
            send_mail(subject, message, "noreply@tuweb.com", [user.email])

            return redirect(reverse("libreria:password_reset_done") + "?sent=true")

        except User.DoesNotExist:
            return redirect(reverse("libreria:password_reset_done") + "?sent=false")

    return render(request, "password_reset.html")


def password_reset_done(request):
    return render(request, "password_reset_done.html")


def password_reset_confirm(request, uidb64, token):
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None

    if user is not None and default_token_generator.check_token(user, token):
        if request.method == "POST":
            form = SetPasswordForm(user, request.POST)
            if form.is_valid():
                form.save()
                return redirect(reverse("libreria:password_reset_complete"))
        else:
            form = SetPasswordForm(user)
        return render(request, "password_reset_confirm.html", {"form": form})
    else:
        return render(request, "password_reset_confirm.html", {"error": "El enlace no es válido o ha expirado."})


def password_reset_complete(request):
    return render(request, "password_reset_complete.html")

#VALIDAR INFORMACION

def validar_datos(request):
    email = request.GET.get('email', None)
    
    errores = {}

    # Validar correo electrónico
    if email and CustomUser.objects.filter(email=email).exists():
        errores['email'] = 'El email ya está en uso.'

    # Retornar los errores (si los hay) o una respuesta de validación exitosa
    return JsonResponse(errores if errores else {'valid': True})

def validate_password(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        password = data.get('password')
        user = request.user
        valid = user.check_password(password)
        return JsonResponse({'valid': valid})
    return JsonResponse({'valid': False})

from io import BytesIO
from datetime import datetime
from django.http import HttpResponse
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from django.contrib.staticfiles import finders
from papeleria.models import Articulo
def draw_table_on_canvas(canvas, doc):
    # Marca de agua
    watermark_path = finders.find('imagen/LOGO.png')
    if watermark_path:
        canvas.saveState()
        canvas.setFillColor(colors.Color(1, 1, 1, alpha=0.3))
        canvas.setStrokeColor(colors.Color(1, 1, 1, alpha=0.3))
        canvas.drawImage(watermark_path,
                         x=(doc.pagesize[0] - 600) / 2,
                         y=(doc.pagesize[1] - 600) / 2,
                         width=600, height=600, mask='auto')
        canvas.restoreState()

def reporte_usuario_pdf(request):
    buffer = BytesIO()

    # Crear documento
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                            leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)

    # Metadatos
    doc.title = "Listado de usuarios CCD"
    doc.author = "CCD"
    doc.subject = "Listado de usuarios"
    doc.creator = "Sistema de Gestión CCD"

    elements = []
    styles = getSampleStyleSheet()

    # Título
    titulo = Paragraph("REPORTE DE USUARIOS", styles["Title"])
    elements.append(titulo)

    # Encabezado empresa
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    encabezado_data = [
        ["GESTOR CCD", "Lista de usuarios", "Correo: gestorccd@gmail.com", f"Fecha: {fecha_actual}"],
        ["Cámara de comercio de Duitama", "Nit: 123456789", "(Correo de la camara)", "Teléfono: (tel. camara)"],
    ]
    tabla_encabezado = Table(encabezado_data, colWidths=[180, 180, 180, 180])
    estilo_encabezado = TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    tabla_encabezado.setStyle(estilo_encabezado)
    elements.append(tabla_encabezado)

    # Tabla usuario
    usuario = request.user
    data_usuario = [["Usuario:", "Email:", "Rol:", "Cargo:"]]
    data_usuario.append([
        usuario.username,
        usuario.email,
        getattr(usuario, 'role', 'No definido'),
        getattr(usuario, 'cargo', 'No definido'),
    ])
    table_usuario = Table(data_usuario, colWidths=[180, 180, 180, 180])
    style_usuario = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table_usuario.setStyle(style_usuario)
    elements.append(table_usuario)

    # Tabla artículos
    data_usuarios = [["ID", "Usuario", "Rol", "correo", "Cargo", "Área", "Estado"]]
    estado = "Activo" if usuario.is_active else "Inactivo"
    for usuario in CustomUser.objects.all():
        data_usuarios.append([
            usuario.id,
            usuario.username,
            usuario.role,
            usuario.email,
            usuario.cargo,
            usuario.area,
             "Activo" if usuario.is_active else "Inactivo"
        ])
    tabla_articulos = Table(data_usuarios, colWidths=[20, 100, 80, 180, 220, 80, 40])
    style_articulos = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    tabla_articulos.setStyle(style_articulos)
    elements.append(tabla_articulos)

    # Construir documento con marca de agua
    doc.build(elements, onFirstPage=draw_table_on_canvas, onLaterPages=draw_table_on_canvas)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Lista de usuarios Gestor CCD.pdf"'
    return response

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl import Workbook
from openpyxl.drawing.image import Image

def reporte_usuario_excel(request):
    # Crear un nuevo libro de trabajo de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Artículos"

    # Ajustar el ancho de las columnas A y B para que quepa el logo
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25

    # Ajustar la altura de la fila 1 para el logo
    ws.row_dimensions[1].height = 90

    # Estilo de borde delgado
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
# Agregar el logo
    logo_path = finders.find('imagen/logo.png')  # Ajustar ruta según ubicación real
    if logo_path:
        img = Image(logo_path)
        img.height = 80
        img.width = 200
        ws.add_image(img, 'A1')
        ws.merge_cells('A1:B1')

    # Aplicar bordes a las celdas del logo (A1 y B1)
    for col in ['A', 'B']:
        cell = ws[f'{col}1']
        cell.border = border


    # Título principal
    ws.merge_cells('C1:G1')
    ws['C1'] = "GESTOR CCD"
    ws['C1'].font = Font(size=24, bold=True)
    ws['C1'].alignment = Alignment(horizontal='center', vertical='center')

    # Aplicar borde a las celdas del título
    for col in ['C', 'D', 'E', 'F', 'G']:
        cell = ws[f'{col}1']
        cell.border = border

   # Subtítulo
    ws.merge_cells('A2:G2')
    ws['A2'] = "Listado de usuarios Gestor CCD"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

# Aplicar borde a las celdas del subtítulo
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        cell = ws[f'{col}2']
        cell.border = border

    # Encabezados de la tabla
    headers = ["ID", "Usuario", "Rol", "Correo", "Cargo", "Área", "Estado"]
    ws.append(headers)

    # Estilo para los encabezados
    header_fill = PatternFill(start_color="FF0056B3", end_color="FF0056B3", fill_type="solid")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Agregar datos de los usuarios
    for usuario in CustomUser.objects.all():
        estado = "Activo" if usuario.is_active else "Inactivo"
        ws.append([
            usuario.id,
            usuario.username,
            usuario.role,
            usuario.email,
            usuario.cargo,
            usuario.area,
            estado
        ])

    # Ajustar el ancho de las columnas
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15


    # Aplicar estilo a las filas de datos
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

    # Ajustar la altura de las filas del encabezado y subtítulo
    ws.row_dimensions[1].height = 60
    ws.row_dimensions[2].height = 30

    # Preparar la respuesta para descargar el archivo
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Lista de usuarios Gestor CCD.xlsx"'

    # Guardar el archivo en la respuesta
    wb.save(response)
    return response
#GRAFICAS
def graficas_usuarios_activos(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Estadísticas', 'url': reverse('libreria:index_estadistica')}, 
        {'name': 'Gráfico de usuarios activos', 'url': reverse('libreria:graficas_usuarios_activos')}, 
    ]
    
    activos = CustomUser.objects.filter(is_active=True).count()
    inactivos = CustomUser.objects.filter(is_active=False).count()

    nombres = ['Activos', 'Inactivos']
    cantidades = [activos, inactivos]

    return render(request, 'estadisticas/grafica_usuarios.html', {
        'nombres': nombres,
        'cantidades': cantidades,
        'breadcrumbs': breadcrumbs
    })