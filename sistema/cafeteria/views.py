# Django imports
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required
from django.utils.timezone import localtime
from django.contrib import messages
from django.utils.timezone import make_aware
from django.urls import reverse
from django.core.paginator import Paginator
from django.core.mail import send_mail
from django.conf import settings
from django.db.models import Q, Sum, Count
from django.contrib.auth import update_session_auth_hash
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.http import HttpResponse
from django.contrib.staticfiles import finders
from django.db import transaction
from django.utils import timezone
from reportlab.lib.enums import TA_CENTER
from datetime import timedelta
from reportlab.platypus import Spacer
from django.utils import timezone
from django.template.loader import render_to_string
from django.core.mail import EmailMultiAlternatives
from libreria.forms import CustomPasswordChangeForm
from reportlab.lib.styles import ParagraphStyle
# Forms
from cafeteria.forms import LoginForm, ProductoForm, ProductosEditForm, PedidoProductoForm ,  DevolucionFormCaf

# Models
from libreria.models import CustomUser
from .models import Productos, PedidoProducto, Pedido , DevolucionCaf

# Standard libraries
from datetime import datetime
from io import BytesIO

# ReportLab for PDF generation
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

# OpenPyXL for Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def index_caf(request):
    return render(request, 'index_caf/index_caf.html')
# Create your views here.
def login_cafeteria(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']
            user = authenticate(request, email=email, password=password)

            if user is not None:
                if user.module == 'Cafeteria':  # Verifica que el usuario pertenece a Cafetería
                    login(request, user)
                    messages.success(request, "Sesión iniciada correctamente en Cafetería.")
                    return redirect('cafeteria:index_caf')
                else:
                    messages.error(request, "No tienes acceso a este módulo.")
            else:
                messages.error(request, 'Usuario o contraseña incorrectos.')
    else:
        form = LoginForm()

    return render(request, 'login_caf/login_caf.html', {'form': form})

# CERRAR SESIÓN

from .models import Productos  # Asegúrate de que la ruta sea correcta
def error_404_view(request, exception):
    return render(request, 'acceso_denegado.html', status=404)
def timeouterror(request):
    try:
        # Simulación de una operación que puede causar un TimeoutError
        # Aquí va tu lógica real, como una conexión a red, base de datos externa, etc.
        raise TimeoutError("Error de tiempo de espera")  # Simulación

        # Si no ocurre error, puedes devolver otro template si lo deseas
        return render(request, 'exito.html')

    except TimeoutError:
        # Solo captura TimeoutError y redirige a lan_error.html
        return render(request, 'lan_error.html')
@login_required(login_url='/acceso_denegado/')
def index_caf(request):
    es_cafeteria = True

    context = {
        'es_cafeteria': es_cafeteria,
    }
    breadcrumbs = [
        {'name': 'Inicio cafeteria', 'url': '/index_caf'},
    ]
    return render(request, 'index_caf/index_caf.html', {'breadcrumbs': breadcrumbs, **context})

# Create your views here.
def ver_usuario_caf(request, id):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_caf'},
        {'name': 'Ver usuario', 'url': reverse('cafeteria:ver_usuario_caf', kwargs={'id': id})},
    ]
    usuario = get_object_or_404(CustomUser, id=id)
    return render(request, 'usuario_caf/ver_perfil_caf.html', {'usuario': usuario, 'breadcrumbs': breadcrumbs})

def login_cafeteria(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')

        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            user = None

        if user is not None:
            if not user.is_active:
                messages.info(request, "Tu usuario está desactivado. Comunícate con el administrador.")
            else:
                user_auth = authenticate(request, email=email, password=password)
                if user_auth is not None and getattr(user_auth, 'acceso_caf', False):
                    login(request, user_auth)
                    return redirect('cafeteria:index_caf')
                else:
                    messages.error(request, "No tienes permiso para acceder a este módulo.")
        else:
            messages.error(request, "Credenciales inválidas.")

    return render(request, 'login_caf/login_caf.html')
def crear_producto(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': reverse('cafeteria:index_caf')},
        {'name': 'Crear producto', 'url': reverse('cafeteria:crear_producto')},
    ]

    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            producto = form.save(commit=False)
            producto.registrado_por = request.user
            producto.save()
            messages.success(request, '¡Producto creado exitosamente!')
            return redirect('cafeteria:listar_productos')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = ProductoForm()

    return render(request, 'productos/crear_producto.html', {
        'form': form,
        'breadcrumbs': breadcrumbs
    })
def listar_productos(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': reverse('cafeteria:index_caf')},
        {'name': 'Listado de cafeteria y aseo', 'url': reverse('cafeteria:listar_productos')},
    ]

    query = request.GET.get('q', '').strip()
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    productos = Productos.objects.select_related('registrado_por').all()

    # Búsqueda por texto
    if query:
        productos = productos.filter(
            Q(nombre__icontains=query) |
            Q(marca__icontains=query) |
            Q(presentacion__icontains=query) |
            Q(unidad_medida__icontains=query) |
            Q(precio__icontains=query) |
            Q(proveedor__icontains=query) |
            Q(cantidad__icontains=query) |
            Q(registrado_por__username__icontains=query) |
            Q(id__icontains=query)
        )

    # Filtrado por rango de fechas con timedelta
    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            productos = productos.filter(fecha_registro__gte=fecha_inicio)
        except ValueError:
            productos = Productos.objects.none()

    if fecha_fin_str:
        try:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date() + timedelta(days=1)
            productos = productos.filter(fecha_registro__lt=fecha_fin)
        except ValueError:
            productos = Productos.objects.none()

    # Paginación
    paginator = Paginator(productos, 4)
    page_number = request.GET.get('page')
    productos_page = paginator.get_page(page_number)

    # Build query parameters for pagination
    query_params = ''
    if query:
        query_params += f'&q={query}'
    if fecha_inicio_str:
        query_params += f'&fecha_inicio={fecha_inicio_str}'
    if fecha_fin_str:
        query_params += f'&fecha_fin={fecha_fin_str}'

    return render(request, 'productos/listar_productos.html', {
        'productos': productos_page,
        'breadcrumbs': breadcrumbs,
        'query_params': query_params,
        'current_query': query,
        'current_fecha_inicio': fecha_inicio_str,
        'current_fecha_fin': fecha_fin_str,
    })
def eliminar_producto(request, id):
    producto = get_object_or_404(Productos, id=id)
    if request.method == 'POST':
        producto.delete()
        messages.success(request, "Producto eliminado correctamente.")
        return redirect('cafeteria:listar_productos')
    # En caso de que sea GET u otro método, redirige o muestra algo
    return redirect('cafeteria:listar_productos')
# CERRAR SESIÓN

def editar_producto(request, producto_id):
    producto = get_object_or_404(Productos, id=producto_id)
    
    breadcrumbs = [
        {'name': 'Inicio', 'url': reverse('cafeteria:index_caf')},
        {'name': 'Listado de cafetería y aseo', 'url': reverse('cafeteria:listar_productos')},
        {'name': 'Editar producto', 'url': reverse('cafeteria:editar_producto', args=[producto_id])},  # ✅ Aquí se corrige
    ]

    if request.method == 'POST':
        form = ProductosEditForm(request.POST, instance=producto)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Producto actualizado correctamente.")
            return redirect('cafeteria:listar_productos')
    else:
        form = ProductosEditForm(instance=producto)

    return render(request, 'productos/editar_producto.html', {
        'form': form,
        'producto': producto,
        'breadcrumbs': breadcrumbs,
    })
@login_required(login_url='/acceso_denegado/')

def logout_caf(request):
    logout(request)
    messages.success(request, "Has cerrado sesión correctamente.")
    return redirect(reverse('libreria:inicio'))

User = get_user_model()
def wrap_text(text, max_len=20):
    parts = [text[i:i+max_len] for i in range(0, len(text), max_len)]
    for i in range(len(parts) - 1):
        parts[i] += '-'  # Agrega guion al final de todas menos la última
    return '\n'.join(parts)
#PDF DE PRODUCTOS
def obtener_productos(request):
    query = request.GET.get('q')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    productos = Productos.objects.all()  

    if query:
        productos = productos.filter(
            Q(nombre__icontains=query) |
            Q(marca__icontains=query) |
            Q(presentacion__icontains=query) |
            Q(unidad_medida__icontains=query) |
            Q(precio__icontains=query) |
            Q(proveedor__icontains=query) |
            Q(cantidad__icontains=query) |
            Q(registrado_por__username__icontains=query) |
            Q(id__icontains=query)
        )

    if fecha_inicio and fecha_fin:
        productos = productos.filter(fecha_registro__range=[fecha_inicio, fecha_fin])

    return productos
def draw_table_on_canvas(canvas, doc):
    # Marca de agua
    watermark_path = finders.find('imagen/LOGO.png')
    if watermark_path:
        canvas.saveState()
        canvas.setFillColor(colors.Color(1, 1, 1, alpha=0.3))
        canvas.setStrokeColor(colors.Color(1, 1, 1, alpha=0.3))
        canvas.drawImage(watermark_path,
                         x=(doc.pagesize[0] - 600) / 2,
                         y=(doc.pagesize[1] - 600) / 2,
                         width=600, height=600, mask='auto')
        canvas.restoreState()
def reporte_productos_pdf(request):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                            leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)

    doc.title = "Listado de productos de cafeteria CCD"
    doc.author = "Gestor CCD"
    doc.subject = "Listado de productos de cafeteria CCD"
    doc.creator = "Sistema de Gestión CCD"

    elements = []
    styles = getSampleStyleSheet()

    # Título
    titulo = Paragraph("REPORTE DE CCAFETERIA CCD ", styles["Title"])
    elements.append(titulo)

    # Encabezado empresa
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    encabezado_data = [
        ["GESTOR CCD", "Lista de artículos", "Correo:", f"Fecha: {fecha_actual}"],
        ["Cámara de comercio de Duitama", "Nit: 123456789", "contacto@gestorccd.com", "Teléfono: (123) 456-7890"],
    ]
    tabla_encabezado = Table(encabezado_data, colWidths=[180, 180, 180, 180])
    estilo_encabezado = TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    tabla_encabezado.setStyle(estilo_encabezado)
    elements.append(tabla_encabezado)

    # Tabla usuario
    usuario = request.user
    data_usuario = [["Usuario", "Email", "Rol", "Cargo"]]

    if usuario.is_authenticated:
        data_usuario.append([
            usuario.username,
            usuario.email,
            getattr(usuario, 'role', 'No definido'),
            getattr(usuario, 'cargo', 'No definido'),
        ])
    else:
        data_usuario.append(["Invitado", "-", "-", "-"])

    table_usuario = Table(data_usuario, colWidths=[180, 180, 180, 180])
    style_usuario = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table_usuario.setStyle(style_usuario)
    elements.append(table_usuario)

    # Obtener productos filtrados
    usuarios_filtrados = obtener_productos(request)

    if not usuarios_filtrados.exists():
        centered_style = ParagraphStyle(
            name="CenteredNormal",
            parent=styles["Normal"],
            alignment=TA_CENTER,
        )
        no_results = Paragraph("No se encontraron productos.", centered_style)
        elements.append(no_results)
    else:
        # Tabla artículos
        data_productos = [["ID", "Nombre", "Marca", "Precio", "Cantidad", "Unidad de medida", "Proveedor", "Presentación"]]
        for producto in usuarios_filtrados:
            data_productos.append([
                wrap_text(str(producto.id)),
                wrap_text(producto.nombre),
                wrap_text(producto.marca),
                wrap_text("{:,}".format(producto.precio)),
                wrap_text(str(producto.cantidad)),
                wrap_text(str(producto.unidad_medida)),
                wrap_text(producto.proveedor),
                wrap_text(producto.presentacion),
            ])

        tabla_productos = Table(data_productos, colWidths=[30, 100, 100, 70, 100, 110, 105])
        style_productos = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        tabla_productos.setStyle(style_productos)
        elements.append(tabla_productos)

    # Construir el PDF
    doc.build(elements, onFirstPage=draw_table_on_canvas, onLaterPages=draw_table_on_canvas)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Listado de cafeteria CCD.pdf"'
    return response
def reporte_productos_excel(request):
    # Parámetros de búsqueda
    q = request.GET.get('q', '').strip()
    fecha_inicio = request.GET.get('fecha_inicio', '').strip()
    fecha_fin = request.GET.get('fecha_fin', '').strip()
    marca = request.GET.get('marca', '').strip()

    # Queryset base
    productos = Productos.objects.all()

    # Aplicar filtros
    if q:
        productos = productos.filter(nombre__icontains=q)
    if fecha_inicio:
        try:
            productos = productos.filter(fecha_registro__gte=datetime.strptime(fecha_inicio, '%Y-%m-%d'))
        except ValueError:
            pass
    if fecha_fin:
        try:
            productos = productos.filter(fecha_registro__lte=datetime.strptime(fecha_fin, '%Y-%m-%d'))
        except ValueError:
            pass
    if marca:
        productos = productos.filter(marca__icontains=marca)

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Listado de cafeteria CCD"

    # Borde común
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Tamaño columnas
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 23
    ws.column_dimensions['H'].width = 23

    ws.row_dimensions[1].height = 60
    ws.row_dimensions[2].height = 30

    # Título
    ws.merge_cells('A1:H1')
    ws['A1'] = "GESTOR CCD"
    ws['A1'].font = Font(size=24, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Subtítulo
    ws.merge_cells('A2:H2')
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    ws['A2'] = f"Listado de cafeteria {fecha_actual}"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    # Borde A1:H2
    for row in [1, 2]:
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f"{col}{row}"].border = border
            ws[f"{col}{row}"].alignment = Alignment(horizontal='center', vertical='center')

    # Encabezados
    headers = ["ID", "Nombre", "Marca", "Precio", "Cantidad", "Unidad de medida", "Proveedor", "Presentación"]
    ws.append(headers)
    header_fill = PatternFill(start_color="FF0056B3", end_color="FF0056B3", fill_type="solid")

    # Estilo encabezados fila 3
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        cell = ws[f"{col}3"]
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Aplicar filtro a A3:H3
    ws.auto_filter.ref = "A3:H3"

    # Agregar datos o mostrar mensaje si no hay productos
    if not productos.exists():
        ws.merge_cells('A4:H4')
        cell = ws['A4']
        cell.value = "No se encontraron productos con los filtros aplicados."
        cell.font = Font(italic=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    else:
        for producto in productos:
            ws.append([
                str(producto.id),
                producto.nombre,
                producto.marca,
                "${:,}".format(producto.precio),
                str(producto.cantidad),
                str(producto.unidad_medida),
                producto.proveedor,
                producto.presentacion,
            ])

        # Estilo de celdas de datos
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=8):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Exportar
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Listado de cafeteria CCD.xlsx"'
    wb.save(response)
    return response

def lista_stock_bajo(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': reverse('cafeteria:index_caf')},
        {'name': 'Bajo Stock', 'url': reverse('cafeteria:lista_bajo_stock')},
    ]
    
    query = request.GET.get('q', '').strip()
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    # Filtrar artículos con cantidad menor a 10
    productos = Productos.objects.filter(cantidad__lt=10)

    # Filtro por búsqueda de texto
    if query:
        productos = productos.filter(
            Q(nombre__icontains=query) |
            Q(marca__icontains=query) |
            Q(presentacion__icontains=query) |
            Q(unidad_medida__icontains=query) |
            Q(precio__icontains=query) |
            Q(proveedor__icontains=query) |
            Q(cantidad__icontains=query) |
            Q(registrado_por__username__icontains=query) |
            Q(id__icontains=query)
        )

    # Filtro por fechas con manejo de errores
    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            productos = productos.filter(fecha_registro__gte=fecha_inicio)
        except ValueError:
            productos = Productos.objects.none()

    if fecha_fin_str:
        try:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date() + timedelta(days=1)
            productos = productos.filter(fecha_registro__lt=fecha_fin)
        except ValueError:
            productos = Productos.objects.none()

    bajo_stock = productos.exists()
    nombres_bajo_stock = [art.nombre for art in productos]

    # Paginación
    paginator = Paginator(productos, 4)
    page_number = request.GET.get('page')
    productos_page = paginator.get_page(page_number)

    # Construir parámetros de búsqueda
    query_params = ''
    if query:
        query_params += f'&q={query}'
    if fecha_inicio_str:
        query_params += f'&fecha_inicio={fecha_inicio_str}'
    if fecha_fin_str:
        query_params += f'&fecha_fin={fecha_fin_str}'

    return render(request, 'productos/bajo_stock.html', {
        'productos': productos_page,
        'bajo_stock': bajo_stock,
        'nombres_bajo_stock': nombres_bajo_stock,
        'breadcrumbs': breadcrumbs,
        'query_params': query_params,
        'current_query': query,
        'current_fecha_inicio': fecha_inicio_str,
        'current_fecha_fin': fecha_fin_str,
    })
def crear_pedido_caf(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': reverse('cafeteria:index_caf')},
        {'name': 'Crear pedido', 'url': reverse('cafeteria:crear_pedido_caf')},
    ]

    if request.method == 'POST':
        try:
            estado = 'Confirmado' if request.user.role == 'Administrador' else 'Pendiente'

            productos_ids = request.POST.getlist('producto')
            cantidades = request.POST.getlist('cantidad')
            lugares = request.POST.getlist('lugar')

            area_usuario = getattr(request.user, 'area', 'No establecido')

            # ✅ Obtener fecha personalizada si está activada
            fecha_personalizada_str = request.POST.get('fecha_personalizada')
            if fecha_personalizada_str:
                try:
                    fecha_personalizada = datetime.strptime(fecha_personalizada_str, '%Y-%m-%dT%H:%M')
                    if timezone.is_naive(fecha_personalizada):
                        fecha_personalizada = make_aware(fecha_personalizada)
                except ValueError:
                    fecha_personalizada = timezone.now()
            else:
                fecha_personalizada = timezone.now()

            # Validar stock antes de crear el pedido
            for producto_id, cantidad in zip(productos_ids, cantidades):
                if not producto_id or not cantidad:
                    continue
                try:
                    producto_id = int(producto_id)
                    cantidad = int(cantidad)
                    producto = Productos.objects.get(id=producto_id)
                    if estado == 'Confirmado' and producto.cantidad < cantidad:
                        messages.error(request, f"No hay suficiente stock para el producto: {producto.nombre}. Stock disponible: {producto.cantidad}.")
                        return redirect('cafeteria:crear_pedido_caf')
                except (ValueError, Productos.DoesNotExist):
                    messages.error(request, "Producto inválido o cantidad no válida.")
                    return redirect('cafeteria:crear_pedido_caf')

            # Crear pedido
            pedido = Pedido.objects.create(
                registrado_por=request.user,
                estado=estado,
                fecha_estado=timezone.now() if estado == 'Confirmado' else None,
                fecha_pedido=fecha_personalizada  # ✅ Se aplica fecha personalizada o automática
            )

            for producto_id, cantidad, lugar in zip(productos_ids, cantidades, lugares):
                try:
                    if not producto_id or not cantidad or not lugar:
                        continue

                    producto_id = int(producto_id)
                    cantidad = int(cantidad)
                    producto = Productos.objects.get(id=producto_id)

                    PedidoProducto.objects.create(
                        pedido=pedido,
                        producto=producto,
                        cantidad=cantidad,
                        lugar=lugar,
                        area=area_usuario,
                    )

                    if estado == 'Confirmado':
                        producto.cantidad -= cantidad
                        producto.save()

                except (ValueError, Productos.DoesNotExist) as e:
                    print(f"Error al procesar producto: {e}")
                    continue

            # Notificar a administradores
            admin_users = User.objects.filter(role='Administrador', is_active=True)
            admin_emails = [admin.email for admin in admin_users if admin.email]

            if admin_emails:
                admin_url = request.build_absolute_uri(reverse('cafeteria:mis_pedidos'))
                context = {
                    'usuario': request.user,
                    'pedido': pedido,
                    'admin_url': admin_url,
                    'productos': PedidoProducto.objects.filter(pedido=pedido),
                    'fecha_pedido': localtime(pedido.fecha_pedido).strftime('%d/%m/%Y %H:%M') if pedido.fecha_pedido else '',
                }

                html_message = render_to_string('pedidos/email_notificacion_pedido_caf.html', context)

                text_message = f"""
Hola Administrador,

Se ha registrado un nuevo pedido en el sistema de Cafetería:

Usuario: {request.user.username}
Rol: {request.user.get_role_display()}
Área: {getattr(request.user, 'area', 'No especificada')}
ID del Pedido: {pedido.id}
Estado: {estado}
Fecha: {localtime(pedido.fecha_pedido).strftime('%d/%m/%Y %H:%M') if pedido.fecha_pedido else 'N/D'}

Detalle de productos:
{chr(10).join([f"- {pp.producto.nombre} x {pp.cantidad} (Lugar: {pp.lugar})" for pp in PedidoProducto.objects.filter(pedido=pedido)])}

Revisar el pedido: {admin_url}

Este es un mensaje automático, por favor no respondas.
"""

                try:
                    subject = f"Nuevo pedido {'confirmado' if estado == 'Confirmado' else 'pendiente'} - ID: {pedido.id}"
                    msg = EmailMultiAlternatives(
                        subject,
                        text_message,
                        settings.DEFAULT_FROM_EMAIL,
                        admin_emails
                    )
                    msg.attach_alternative(html_message, "text/html")
                    msg.send()
                except Exception as e:
                    print(f"Error enviando correo: {e}")
                    messages.warning(request, "El pedido fue creado, pero no se pudo enviar el correo de notificación.")

            messages.success(request, f"Pedido creado exitosamente con estado: {estado}.")
            return redirect('cafeteria:mis_pedidos')

        except Exception as e:
            print(f"Error al crear pedido: {e}")
            messages.error(request, "Ocurrió un error al procesar tu pedido. Por favor intenta nuevamente.")
            return redirect('cafeteria:crear_pedido_caf')

    productos = Productos.objects.all()
    return render(request, 'pedidos/pedidos_caf.html', {
        'productos': productos,
        'breadcrumbs': breadcrumbs
    })
@login_required
def mis_pedidos(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': reverse('cafeteria:index_caf')},
        {'name': 'Mis pedidos', 'url': reverse('cafeteria:mis_pedidos')},
    ]
    
    query = request.GET.get('q', '').strip()
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    # Obtener pedidos del usuario autenticado ordenados por fecha descendente
    pedidos = Pedido.objects.filter(registrado_por=request.user).order_by('-fecha_pedido')

    # Filtro de búsqueda
    if query:
        pedidos = pedidos.filter(
            Q(registrado_por__username__icontains=query) |
            Q(estado__icontains=query) |
            Q(id__icontains=query) |
            Q(productos__area__icontains=query) |
            Q(productos__lugar__icontains=query) |  # ✅ Campo añadido
            Q(productos__producto__nombre__icontains=query)
        ).distinct()

    # Filtro por fecha de inicio con manejo de errores
    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            pedidos = pedidos.filter(fecha_pedido__gte=fecha_inicio)
        except ValueError:
            pedidos = Pedido.objects.none()

    # Filtro por fecha de fin con manejo de errores
    if fecha_fin_str:
        try:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date() + timedelta(days=1)
            pedidos = pedidos.filter(fecha_pedido__lt=fecha_fin)
        except ValueError:
            pedidos = Pedido.objects.none()

    # Paginación
    paginator = Paginator(pedidos, 4)
    page_number = request.GET.get('page')
    pedidos_page = paginator.get_page(page_number)

    # Agregar lugares únicos a cada pedido
    for pedido in pedidos_page:
        lugares = pedido.productos.values_list('lugar', flat=True)
        lugares_unicos = list(set(lugares))
        pedido.lugares_unicos = lugares_unicos

    # Construir parámetros de consulta para paginación
    query_params = ''
    if query:
        query_params += f'&q={query}'
    if fecha_inicio_str:
        query_params += f'&fecha_inicio={fecha_inicio_str}'
    if fecha_fin_str:
        query_params += f'&fecha_fin={fecha_fin_str}'

    return render(request, 'pedidos/mis_pedidos_caf.html', {
        'pedidos': pedidos_page,
        'breadcrumbs': breadcrumbs,
        'query_params': query_params,
        'current_query': query,
        'current_fecha_inicio': fecha_inicio_str,
        'current_fecha_fin': fecha_fin_str,
    })
from django.utils.html import escape
@csrf_exempt
@require_POST
def cambiar_estado_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    nuevo_estado = request.POST.get('estado')

    if not nuevo_estado:
        messages.error(request, 'No se especificó un nuevo estado.')
        return redirect('cafeteria:pedidos_pendientes')

    if nuevo_estado == 'Confirmado':
        try:
            with transaction.atomic():
                productos_pedido = pedido.productos.select_related('producto').all()
                productos_sin_stock = []

                for item in productos_pedido:
                    producto = item.producto
                    if producto.cantidad < item.cantidad:
                        productos_sin_stock.append(
                            f"{producto.nombre} (Solicitados: {item.cantidad}, Disponibles: {producto.cantidad})"
                        )

                if productos_sin_stock:
                    pedido.estado = 'Cancelado'
                    pedido.fecha_estado = timezone.now()
                    pedido.save()
                    messages.error(request, f"Pedido cancelado. Stock insuficiente: {', '.join(productos_sin_stock)}")
                    return redirect('cafeteria:pedidos_pendientes')

                for item in productos_pedido:
                    producto = item.producto
                    producto.cantidad -= item.cantidad
                    producto.save()

                pedido.estado = 'Confirmado'
                pedido.fecha_estado = timezone.now()
                pedido.save()

                messages.success(request, 'Pedido confirmado correctamente.')

        except Exception as e:
            messages.error(request, f'Error al confirmar el pedido: {str(e)}')
            return redirect('cafeteria:pedidos_pendientes')

    else:
        pedido.estado = nuevo_estado
        pedido.fecha_estado = timezone.now()
        pedido.save()
        messages.success(request, f'Estado actualizado a {nuevo_estado}.')

    usuario = pedido.registrado_por
    if usuario and usuario.email:
        productos_lista = "\n".join(
            f"<li class='articulo-item'>{escape(item.cantidad)} {escape(item.producto.nombre)}</li>"
            for item in pedido.productos.all()
        )

        html_content = f"""
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <title>Estado de Pedido Actualizado</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            line-height: 1.6;
            color: #333333;
            max-width: 600px;
            margin: 0 auto;
            padding: 20px;
            background-color: #f5f5f5;
        }}
        .email-container {{
            background-color: #ffffff;
            border-radius: 8px;
            box-shadow: 0 2px 10px rgba(0, 0, 0, 0.1);
            overflow: hidden;
        }}
        .header {{
            color: white;
            padding: 20px;
            text-align: center;
        }}
        .logo {{
            max-width: 180px;
            height: auto;
            margin-bottom: 15px;
        }}
        .content {{
            padding: 25px;
        }}
        .status {{
            font-size: 18px;
            font-weight: bold;
            margin: 15px 0;
            padding: 10px;
            border-radius: 5px;
            text-align: center;
        }}
        .status.Confirmado {{
            color: #155724;
        }}
        .status.Cancelado {{
            background-color: #f8d7da;
            color: #721c24;
        }}
        .status.Pendiente {{
            color: #856404;
        }}
        .info-box {{
            background-color: #f8f9fa;
            border-left: 4px solid #007bff;
            padding: 15px;
            margin: 20px 0;
        }}
        .info-label {{
            font-weight: bold;
            color: #495057;
            display: inline-block;
            min-width: 150px;
        }}
        .articulos-list {{
            margin: 15px 0;
            padding-left: 20px;
        }}
        .articulo-item {{
            margin-bottom: 8px;
        }}
        .footer {{
            text-align: center;
            padding: 15px;
            font-size: 12px;
            color: #6c757d;
            border-top: 1px solid #e9ecef;
            background-color: #f8f9fa;
        }}
    </style>
</head>
<body>
    <div class="email-container">
        <div class="header">
            <img src="https://ccduitama.org.co/wp-content/uploads/2021/05/LOGOCCD-TRANSPARENCIA.png" alt="Logo CCD" class="logo">
            <h2>Actualización de Estado de Pedido</h2>
        </div>
        
        <div class="content">
            <p>Hola <strong>{escape(usuario.username)}</strong>,</p>
            <p>Te informamos que el estado de tu pedido del módulo de cafeteria ha sido actualizado:</p>
            
            <div class="info-box">
                <p><span class="info-label">Número de Pedido:</span> #{pedido.id}</p>
                <p><span class="info-label">Fecha de Actualización:</span> {pedido.fecha_estado.strftime('%d/%m/%Y %H:%M')}</p>
                <p><span class="info-label">Nuevo Estado:</span> <span class="status {pedido.estado}">{pedido.estado.upper()}</span></p>

                <h3>Detalle del Pedido:</h3>
                <ul class="articulos-list">
                    {productos_lista}
                </ul>
            </div>
            
            <p>Para más información, puedes acceder al sistema de gestión de pedidos.</p>
        </div>
        
        <div class="footer">
            <p>Este es un mensaje automático, por favor no respondas a este correo.</p>
            <p>© {timezone.now().year} Gestor Cafetería - Todos los derechos reservados</p>
        </div>
    </div>
</body>
</html>
        """

        text_content = f"""
Actualización de Estado de Pedido

Hola {usuario.username},

Tu pedido #{pedido.id} ha sido actualizado a {pedido.estado.upper()}.

Fecha de actualización: {pedido.fecha_estado.strftime('%d/%m/%Y %H:%M')}

Detalle del Pedido:
{"".join([f"- {item.cantidad} x {item.producto.nombre}\n" for item in pedido.productos.all()])}

Gracias por usar nuestro sistema.
© {timezone.now().year} Gestor Cafetería
        """

        try:
            send_mail(
                subject=f'Pedido #{pedido.id} - Estado actualizado a {pedido.estado.upper()}',
                message=text_content,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[usuario.email],
                html_message=html_content
            )
        except Exception as e:
            print(f"Error enviando email: {str(e)}")

    return redirect('cafeteria:pedidos_pendientes')

    messages.error(request, 'No se pudo actualizar el estado.')
    return redirect('cafeteria:pedidos_pendientes')

def pedidos_pendientes(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': reverse('cafeteria:index_caf')},
        {'name': 'Pedidos pendientes', 'url': reverse('cafeteria:pedidos_pendientes')},
    ]
    
    query = request.GET.get('q', '').strip()
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    # Obtener pedidos pendientes ordenados por fecha descendente
    pedidos = Pedido.objects.filter(estado='Pendiente').order_by('-fecha_pedido')

    # Filtro de búsqueda de texto
    if query:
        pedidos = pedidos.filter(
            Q(registrado_por__username__icontains=query) |
            Q(estado__icontains=query) |
            Q(id__icontains=query) |
            Q(productos__area__icontains=query) |
            Q(productos__lugar__icontains=query) |  # ✅ Campo añadido
            Q(productos__producto__nombre__icontains=query)
        ).distinct()

    # Filtro por fecha de inicio
    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            pedidos = pedidos.filter(fecha_pedido__gte=fecha_inicio)
        except ValueError:
            pass

    # Filtro por fecha de fin
    if fecha_fin_str:
        try:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date() + timedelta(days=1)
            pedidos = pedidos.filter(fecha_pedido__lt=fecha_fin)
        except ValueError:
            pass

    # Paginación
    paginator = Paginator(pedidos, 4)
    page_number = request.GET.get('page')
    pedidos_page = paginator.get_page(page_number)

    # Agregar lugares únicos a cada pedido
    for pedido in pedidos_page:
        lugares = pedido.productos.values_list('lugar', flat=True)
        lugares_unicos = list(set(lugares))
        pedido.lugares_unicos = lugares_unicos

    # Construir parámetros para mantener filtros en paginación
    query_params = ''
    if query:
        query_params += f'&q={query}'
    if fecha_inicio_str:
        query_params += f'&fecha_inicio={fecha_inicio_str}'
    if fecha_fin_str:
        query_params += f'&fecha_fin={fecha_fin_str}'

    return render(request, 'pedidos/confirmar_pedido_caf.html', {
        'pedidos': pedidos_page,
        'breadcrumbs': breadcrumbs,
        'query_params': query_params,
        'current_query': query,
        'current_fecha_inicio': fecha_inicio_str,
        'current_fecha_fin': fecha_fin_str,
    })
def listado_pedidos_caf(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': reverse('cafeteria:index_caf')},
        {'name': 'Listado de pedidos', 'url': reverse('cafeteria:lista_pedidos_caf')},
    ]

    query = request.GET.get('q', '').strip()
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    # Obtener todos los pedidos ordenados por fecha descendente
    pedidos = Pedido.objects.all().order_by('-fecha_pedido')

    # Filtro de búsqueda de texto
    if query:
        pedidos = pedidos.filter(
            Q(registrado_por__username__icontains=query) |
            Q(estado__icontains=query) |
            Q(id__icontains=query) |
            Q(productos__area__icontains=query) |
            Q(productos__lugar__icontains=query) |  # ✅ Campo añadido
            Q(productos__producto__nombre__icontains=query)
        ).distinct()

    # Filtro por fecha de inicio
    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            pedidos = pedidos.filter(fecha_pedido__gte=fecha_inicio)
        except ValueError:
            pass  # Ignora fechas inválidas

    # Filtro por fecha de fin
    if fecha_fin_str:
        try:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date() + timedelta(days=1)
            pedidos = pedidos.filter(fecha_pedido__lt=fecha_fin)
        except ValueError:
            pass  # Ignora fechas inválidas

    # Paginación
    paginator = Paginator(pedidos, 4)
    page_number = request.GET.get('page')
    pedidos_page = paginator.get_page(page_number)

    # Agregar lugares únicos a cada pedido
    for pedido in pedidos_page:
        lugares = pedido.productos.values_list('lugar', flat=True)
        lugares_unicos = list(set(lugares))
        pedido.lugares_unicos = lugares_unicos

    # Parámetros de consulta para mantener filtros en la paginación
    query_params = ''
    if query:
        query_params += f'&q={query}'
    if fecha_inicio_str:
        query_params += f'&fecha_inicio={fecha_inicio_str}'
    if fecha_fin_str:
        query_params += f'&fecha_fin={fecha_fin_str}'

    return render(request, 'pedidos/lista_pedidos_caf.html', {
        'pedidos': pedidos_page,
        'breadcrumbs': breadcrumbs,
        'query_params': query_params,
        'current_query': query,
        'current_fecha_inicio': fecha_inicio_str,
        'current_fecha_fin': fecha_fin_str,
    })

def get_pedidos_filtrados_caf(request):
    query = request.GET.get('q')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    pedidos = Pedido.objects.all()

    if query:
        pedidos = pedidos.filter(
            Q(registrado_por__username__icontains=query) |
            Q(estado__icontains=query) |
            Q(id__icontains=query) |
            Q(productos__area__icontains=query) |
            Q(productos__lugar__icontains=query) |  # ✅ Campo añadido
            Q(productos__producto__nombre__icontains=query)
        ).distinct()

    if fecha_inicio and fecha_fin:
        pedidos = pedidos.filter(fecha_pedido__range=[fecha_inicio, fecha_fin])

    return pedidos
def wrap_text_p(text, max_len=26
                ):
    parts = [text[i:i+max_len] for i in range(0, len(text), max_len)]
    for i in range(len(parts) - 1):
        parts[i] += '-'  # Agrega guion al final de todas menos la última
    return '\n'.join(parts)

def reporte_pedidos_pdf_caf(request):
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=40,
        rightMargin=40,
        topMargin=40,
        bottomMargin=40
    )

    doc.title = "Listado de pedidos de cafeteria y aseo CCD"
    doc.author = "CCD"
    doc.subject = "Listado de pedidos"
    doc.creator = "Sistema de Gestión CCD"

    elements = []
    styles = getSampleStyleSheet()

    # Título principal
    titulo = Paragraph("REPORTE DE PEDIDOS DE CAFETERIA Y ASEO", styles["Title"])
    elements.append(titulo)

    # Encabezado institucional
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    encabezado_data = [
        ["GESTOR CCD", "Lista de usuarios", "Correo: gestorccd@gmail.com", f"Fecha: {fecha_actual}"],
        ["Cámara de comercio de Duitama", "Nit: 123456789", "(Correo de la camara)", "Teléfono: (tel. camara)"],
    ]
    tabla_encabezado = Table(encabezado_data, colWidths=[180, 180, 180, 180])
    estilo_encabezado = TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    tabla_encabezado.setStyle(estilo_encabezado)
    elements.append(tabla_encabezado)

    # Datos del usuario
    usuario = request.user
    data_usuario = [["Usuario:", "Email:", "Rol:", "Cargo:"]]
    data_usuario.append([
        usuario.username,
        usuario.email,
        getattr(usuario, 'role', 'No definido'),
        getattr(usuario, 'cargo', 'No definido'),
    ])
    table_usuario = Table(data_usuario, colWidths=[180, 180, 180, 180])
    style_usuario = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table_usuario.setStyle(style_usuario)
    elements.append(table_usuario)

    # Obtener filtros
    q = request.GET.get('q', '')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    # Fetch pedidos with prefetch for productos and producto
    pedidos = Pedido.objects.prefetch_related('productos__producto').all()

    if q:
        pedidos = pedidos.filter(
            Q(id__icontains=q) |
            Q(productos__producto__nombre__icontains=q) |
            Q(registrado_por__username__icontains=q) |
            Q(estado__icontains=q)
        ).distinct()

    if fecha_inicio:
        try:
            pedidos = pedidos.filter(fecha_pedido__gte=datetime.strptime(fecha_inicio, '%Y-%m-%d'))
        except ValueError:
            pass  # Handle invalid date format gracefully
    if fecha_fin:
        try:
            pedidos = pedidos.filter(fecha_pedido__lte=datetime.strptime(fecha_fin, '%Y-%m-%d'))
        except ValueError:
            pass  # Handle invalid date format gracefully

    if not pedidos.exists():
        centered_style = ParagraphStyle(
            name="CenteredNormal",
            parent=styles["Normal"],
            alignment=TA_CENTER,
        )
        no_results = Paragraph("No se encontraron pedidos.", centered_style)
        elements.append(no_results)
    else:
        # Encabezado de la tabla
        data_pedidos = [["ID Pedido", "Fecha", "Estado", "Registrado Por", "Productos", "Área"]]

        # Filas de pedidos
        for pedido in pedidos:
            try:
                productos = pedido.productos.all()
                if productos.exists():
                    productos_raw = ", ".join([
                        f"{pa.producto.nombre} x {pa.cantidad}"
                        for pa in productos
                    ])
                    areas = set(pa.area for pa in productos if pa.area and pa.area != 'No establecido')
                    area_raw = ", ".join(areas) if areas else 'No establecido'
                else:
                    productos_raw = 'Sin productos'
                    area_raw = 'Sin área'

                data_pedidos.append([
                    wrap_text_p(str(pedido.id)),
                    wrap_text_p(pedido.fecha_pedido.strftime('%d-%m-%Y')),
                    wrap_text_p(pedido.get_estado_display()),
                    wrap_text_p(pedido.registrado_por.username if pedido.registrado_por else 'No definido'),
                    wrap_text_p(productos_raw),
                    wrap_text_p(area_raw)
                ])
            except Exception as e:
                print(f"Error processing pedido {pedido.id}: {str(e)}")
                data_pedidos.append([
                    wrap_text_p(str(pedido.id)),
                    wrap_text_p(pedido.fecha_pedido.strftime('%d-%m-%Y')),
                    wrap_text_p(pedido.get_estado_display()),
                    wrap_text_p(pedido.registrado_por.username if pedido.registrado_por else 'No definido'),
                    wrap_text_p('Error al cargar productos'),
                    wrap_text_p('Error al cargar área')
                ])

        # Crear la tabla de pedidos
        tabla_productos = Table(data_pedidos, colWidths=[60, 100, 100, 160, 200, 100])
        style_productos = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        tabla_productos.setStyle(style_productos)
        elements.append(tabla_productos)

    # Construir el PDF
    doc.build(elements, onFirstPage=draw_table_on_canvas, onLaterPages=draw_table_on_canvas)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Lista_de_pedidos_Gestor_CCD.pdf"'
    return response
def reporte_pedidos_excel_caf(request):
    pedidos = Pedido.objects.all().order_by('-fecha_pedido')

    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos CCD"

    columnas_anchos = {
        'A': 10,
        'B': 15,
        'C': 15,
        'D': 25,
        'E': 50,
        'F': 30
    }
    for col, width in columnas_anchos.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 50
    ws.row_dimensions[2].height = 30

    ws.merge_cells('A1:F1')
    ws['A1'] = "GESTOR CCD"
    ws['A1'].font = Font(size=24, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:F2')
    ws['A2'] = "Listado de Pedidos"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=6):
        for cell in row:
            cell.border = thin_border

    headers = ['ID', 'Fecha', 'Estado', 'Registrado Por', 'Productos', 'Áreas']
    ws.append(headers)

    header_fill = PatternFill(start_color="FF0056B3", end_color="FF0056B3", fill_type="solid")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.auto_filter.ref = "A3:F3"

    # Agregar datos o mensaje si no hay pedidos
    if not pedidos.exists():
        ws.merge_cells('A4:F4')
        cell = ws['A4']
        cell.value = "No se encontraron pedidos."
        cell.font = Font(italic=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws.row_dimensions[4].height = 40
    else:
        for pedido in pedidos:
            productos_raw = ", ".join([
                f"{pa.producto.nombre} x {pa.cantidad} ({pa.lugar})"
                for pa in pedido.productos.all()
            ]) or 'Sin artículos'

            areas_raw = ", ".join(set([
                str(pa.area) for pa in pedido.productos.all()
            ])) or 'Sin área'

            usuario = pedido.registrado_por.username if pedido.registrado_por else 'No definido'

            ws.append([
                pedido.id,
                pedido.fecha_pedido.strftime('%Y-%m-%d'),
                pedido.get_estado_display(),
                usuario,
                productos_raw,
                areas_raw,
            ])

        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=6):
            for i, cell in enumerate(row, start=1):
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=True if i in [4, 5, 6] else False
                )

        for i in range(4, ws.max_row + 1):
            ws.row_dimensions[i].height = 60

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Reporte_pedidos.xlsx"'
    wb.save(response)
    return response

#ESTADISTICAS 
def index_estadistica_caf(request):
    breadcrumbs = [
    {'name': 'Inicio', 'url': '/index_caf'},
    {'name': 'Estadisticas', 'url': reverse('cafeteria:index_estadistica_caf')}, 
]

    return render(request, 'estadisticas_caf/index_estadistica_caf.html', {'breadcrumbs': breadcrumbs})
#GRAFIC DE CANTIDAD DE PRODUCTOS
def graficas_productos(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_caf'},
        {'name': 'Estadisticas', 'url': reverse('cafeteria:index_estadistica_caf')}, 
        {'name': 'Grafico de productos', 'url': reverse('cafeteria:graficas_productos')}, 
    ]
    
    # Obtener parámetros de fecha
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    productos = Productos.objects.all()
    
    # Aplicar filtros de fecha si existen
    if fecha_inicio:
        productos = productos.filter(fecha_registro__gte=fecha_inicio)
    if fecha_fin:
        # Convertir a datetime y sumar 1 día para incluir todo el día final
        fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d') + timedelta(days=1)
        productos = productos.filter(fecha_registro__lt=fecha_fin_dt)
    
    nombres = [art.nombre for art in productos]
    cantidades = [art.cantidad for art in productos]

    return render(request, 'estadisticas_caf/grafica_productos.html', {
        'nombres': nombres,
        'cantidades': cantidades,
        'breadcrumbs': breadcrumbs,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
    })
#grafica de usuario

def graficas_usuario_caf(request):
    breadcrumbs = [
         {'name': 'Inicio', 'url': '/index_caf'},
        {'name': 'Estadisticas', 'url': reverse('cafeteria:index_estadistica_caf')}, 
        {'name': 'Grafico de usuarios activos e inactivos', 'url': reverse('cafeteria:graficas_usuarios_caf')}, 
    ]
    activos = CustomUser.objects.filter(acceso_caf=True).filter(is_active=True).count()
    inactivos = CustomUser.objects.filter(acceso_caf=True).filter(is_active=False).count()

    nombres = ['Activos', 'Inactivos']
    cantidades = [activos, inactivos] 
    return render(request, 'estadisticas_caf/grafica_usuarios_caf.html', {'nombres':nombres,
                                                                  'cantidades': cantidades,
                                                                  'breadcrumbs': breadcrumbs})

def grafica_pedidos_caf(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_caf'},
        {'name': 'Estadisticas', 'url': reverse('cafeteria:index_estadistica_caf')}, 
        {'name': 'Gráfico de pedidos Administrativa', 'url': reverse('cafeteria:grafica_pedidos_caf')},
    ]

    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    productos = PedidoProducto.objects.select_related('producto')

    if fecha_inicio:
        productos = productos.filter(pedido__fecha_pedido__gte=fecha_inicio)
    if fecha_fin:
        fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d') + timedelta(days=1)
        productos = productos.filter(pedido__fecha_pedido__lt=fecha_fin_dt)

    datos = productos.values('producto__nombre').annotate(
        total_cantidad=Sum('cantidad')
    ).order_by('producto__nombre')

    etiquetas = [item['producto__nombre'] for item in datos]
    cantidades = [item['total_cantidad'] for item in datos]

    return render(request, 'estadisticas_caf/grafico_pedidos_caf.html', {
        'nombres': etiquetas,
        'cantidades': cantidades,
        'breadcrumbs': breadcrumbs,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
    })
def grafica_estado_pedido_caf(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_caf'},
        {'name': 'Estadisticas', 'url': reverse('cafeteria:index_estadistica_caf')}, 
        {'name': 'Gráfico de estado de pedidos', 'url': reverse('cafeteria:grafica_estado_pedido_caf')},
    ]

    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    pedidos = Pedido.objects.all()

    if fecha_inicio:
        pedidos = pedidos.filter(fecha_pedido__gte=fecha_inicio)
    if fecha_fin:
        # Incluir todo el día final sumando 1 día
        fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d') + timedelta(days=1)
        pedidos = pedidos.filter(fecha_pedido__lt=fecha_fin_dt)

    pendientes = pedidos.filter(estado='Pendiente').count()
    confirmados = pedidos.filter(estado='Confirmado').count()
    cancelados = pedidos.filter(estado='Cancelado').count()

    nombres = ['Pendiente', 'Confirmado', 'Cancelado']
    cantidades = [pendientes, confirmados, cancelados]

    return render(request, 'estadisticas_caf/grafica_estado_pendiente_caf.html', {
        'nombres': nombres,
        'cantidades': cantidades,
        'breadcrumbs': breadcrumbs,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin
    })
def grafica_bajo_Stock_caf(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_caf'},
        {'name': 'Estadisticas', 'url': reverse('cafeteria:index_estadistica_caf')}, 
        {'name': 'Grafico de bajo stock', 'url': reverse('cafeteria:grafica_bajoStock_caf')}, 
    ]
    productos = Productos.objects.filter(cantidad__lt=10)
    nombres = [art.nombre for art in productos]
    cantidades = [art.cantidad for art in productos]
    return render(request, 'estadisticas_caf/grafica_bajoStock_caf.html', {
        'nombres': nombres,
        'cantidades': cantidades,
        'breadcrumbs': breadcrumbs
    })

#PDF
def get_pedidos_filtrados_pendientes_caf(request):
    query = request.GET.get('q')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    pedidos = Pedido.objects.filter(estado='Pendiente').order_by('-fecha_pedido')
    if query:
        pedidos = pedidos.filter(
            Q(registrado_por__username__icontains=query) |
            Q(estado__icontains=query) |
            Q(productos__producto__nombre__icontains=query) |
            Q(productos__cantidad__icontains=query) |
            Q(productos__area__icontains=query)
    ).distinct()

    if fecha_inicio and fecha_fin:
        try:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d") + timedelta(days=1)
            pedidos = pedidos.filter(fecha_pedido__gte=fecha_inicio_dt, fecha_pedido__lt=fecha_fin_dt)
        except ValueError:
            pass  # opcional: puedes registrar/loggear el error si las fechas son inválidas

    return pedidos
def reporte_pedidos_pendientes_pdf_caf(request):
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=40,
        rightMargin=40,
        topMargin=40,
        bottomMargin=40
    )

    doc.title = "Pedidos pendientes Cafeteria y Aseo CCD"
    doc.author = "GESTOR CCD"
    doc.subject = "Pedidos pendientes CCD"
    doc.creator = "Sistema de Gestión CCD"

    elements = []
    styles = getSampleStyleSheet()

    # Título principal
    titulo = Paragraph("Reporte de pedidos pendientes cafeteria y aseo", styles["Title"])
    elements.append(titulo)

    # Encabezado institucional
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    encabezado_data = [
        ["GESTOR CCD", "Lista de usuarios", "Correo: gestorccd@gmail.com", f"Fecha: {fecha_actual}"],
        ["Cámara de comercio de Duitama", "Nit: 123456789", "(Correo de la camara)", "Teléfono: (tel. camara)"],
    ]
    tabla_encabezado = Table(encabezado_data, colWidths=[180, 180, 180, 180])
    estilo_encabezado = TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    tabla_encabezado.setStyle(estilo_encabezado)
    elements.append(tabla_encabezado)

    # Datos del usuario
    usuario = request.user
    data_usuario = [["Usuario:", "Email:", "Rol:", "Cargo:"]]
    data_usuario.append([
        usuario.username,
        usuario.email,
        getattr(usuario, 'role', 'No definido'),
        getattr(usuario, 'cargo', 'No definido'),
    ])
    table_usuario = Table(data_usuario, colWidths=[180, 180, 180, 180])
    style_usuario = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table_usuario.setStyle(style_usuario)
    elements.append(table_usuario)

    # Obtener pedidos filtrados
    pedidos = get_pedidos_filtrados_pendientes_caf(request).prefetch_related('productos__producto')

    if not pedidos.exists():
        centered_style = ParagraphStyle(
            name="CenteredNormal",
            parent=styles["Normal"],
            alignment=TA_CENTER
        )
        no_results = Paragraph("No se encontraron pedidos pendientes.", centered_style)
        elements.append(Spacer(1, 20))
        elements.append(no_results)
    else:
        # Encabezado de la tabla
        data_pedidos = [["ID Pedido", "Fecha", "Estado", "Registrado Por", "Artículos", "Área"]]

        # Filas de pedidos
        for pedido in pedidos:
            productos_raw = ", ".join([
                f"{pa.cantidad} {pa.producto.nombre}({pa.lugar})"
                for pa in pedido.productos.all()
            ]) or 'Sin artículos'
            productos_text = wrap_text_p(productos_raw)

            area_raw = pedido.productos.first().area if pedido.productos.exists() else 'Sin área'
            areas_text = wrap_text_p(area_raw)

            data_pedidos.append([
                wrap_text_p(str(pedido.id)),
                wrap_text_p(pedido.fecha_pedido.strftime('%d-%m-%Y')),
                wrap_text_p(pedido.get_estado_display()),
                wrap_text_p(pedido.registrado_por.username if pedido.registrado_por else 'No definido'),
                wrap_text_p(productos_raw),
                wrap_text_p(area_raw)
            ])

        # Crear tabla de pedidos
        tabla_productos = Table(data_pedidos, colWidths=[60, 100, 100, 160, 200, 100])
        style_productos = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        tabla_productos.setStyle(style_productos)
        elements.append(tabla_productos)

    # Construir el PDF
    doc.build(elements, onFirstPage=draw_table_on_canvas, onLaterPages=draw_table_on_canvas)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Pedidos pendientes cafeteria y aseo CCD.pdf"'
    return response

def reporte_pedidos_pendientes_excel_caf(request):
    pedidos = get_pedidos_filtrados_pendientes_caf(request).prefetch_related('productos__producto')

    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos CCD"

    # Configurar anchos de columna
    columnas_anchos = {'A': 10, 'B': 15, 'C': 15, 'D': 25, 'E': 50, 'F': 30}
    for col, width in columnas_anchos.items():
        ws.column_dimensions[col].width = width

    # Alturas para título y subtítulo
    ws.row_dimensions[1].height = 60
    ws.row_dimensions[2].height = 30

    # Título principal
    ws.merge_cells('A1:F1')
    ws['A1'] = "GESTOR CCD"
    ws['A1'].font = Font(size=24, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Subtítulo con fecha
    ws.merge_cells('A2:F2')
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    ws['A2'] = f"Pedidos pendiente cafeteria y aseo - {fecha_actual}"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    # Borde del título y subtítulo
    borde = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=6):
        for cell in row:
            cell.border = borde
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Encabezados
    headers = ['ID', 'Fecha', 'Estado', 'Registrado Por', 'Artículos', 'Áreas']
    ws.append(headers)
    header_fill = PatternFill(start_color="FF0056B3", end_color="FF0056B3", fill_type="solid")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Filtro automático
    ws.auto_filter.ref = "A3:F3"

    # Agregar datos
    if not pedidos.exists():
        ws.merge_cells('A4:F4')
        cell = ws['A4']
        cell.value = "No se encontraron pedidos pendientes."
        cell.font = Font(bold=False)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = borde
    else:
        for pedido in pedidos:
            productos_raw = ", ".join([
                f"{pa.producto.nombre} x {pa.cantidad} ({pa.lugar})"
                for pa in pedido.productos.all()
            ])
            areas_raw = ", ".join(set(str(pa.area) for pa in pedido.productos.all()))  # ← corregido aquí
            usuario = pedido.registrado_por.username if pedido.registrado_por else 'No definido'
            ws.append([
                pedido.id,
                pedido.fecha_pedido.strftime('%Y-%m-%d'),
                pedido.get_estado_display() if hasattr(pedido, 'get_estado_display') else pedido.estado,
                usuario,
                productos_raw or 'Sin artículos',
                areas_raw or 'Sin área',
            ])

    # Estilos de celdas: bordes y alineación
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=6):
        for i, cell in enumerate(row, start=1):
            cell.border = borde
            cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True if i in [4, 5, 6] else False
            )

    # Altura dinámica para filas de contenido
    for i in range(4, ws.max_row + 1):
        ws.row_dimensions[i].height = 60

    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Reporte_pedidos_pendientes.xlsx"'
    wb.save(response)
    return response

def cambiar_contraseña_caf(request):
    breadcrumbs = [
        {'name': 'Inicio cafeteria', 'url': '/index_pap'},
        {'name': 'Cambiar Contraseña cafeteria', 'url': reverse('cafeteria:cambiar_contraseña_caf')},
    ]
    if request.method == 'POST':
        form = CustomPasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            return redirect('libreria:inicio')  # Redirige después de cambiar la contraseña
    else:
        form = CustomPasswordChangeForm(user=request.user)
    return render(request, 'usuario_caf/cambiar_contraseña_caf.html', {'form': form, 'breadcrumbs': breadcrumbs})


def obtener_producto_bajo_stock_caf(request):
    query = request.GET.get('q')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    productos = Productos.objects.filter(cantidad__lt=10)

    if query:
        productos = productos.filter(
            Q(nombre__icontains=query) |
            Q(marca__icontains=query) |
            Q(presentacion__icontains=query) |
            Q(unidad_medida__icontains=query) |
            Q(precio__icontains=query) |
            Q(proveedor__icontains=query) |
            Q(cantidad__icontains=query) |
            Q(registrado_por__username__icontains=query) |
            Q(id__icontains=query)
        )

    if fecha_inicio and fecha_fin:
        productos = productos.filter(fecha_registro__range=[fecha_inicio, fecha_fin])

    return productos

def reporte_producto_bajo_stock_pdf(request):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                            leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)

    doc.title = "Listado de BAJO STOCK CCD"
    doc.author = "CCD"
    doc.subject = "Listado de bajo stock"
    doc.creator = "Sistema de Gestión CCD"

    elements = []
    styles = getSampleStyleSheet()

    # Título
    titulo = Paragraph("REPORTE DE BAJO STOCK", styles["Title"])
    elements.append(titulo)

    # Encabezado empresa
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    encabezado_data = [
        ["GESTOR CCD", "Lista de artículos", "Correo:", f"Fecha: {fecha_actual}"],
        ["Cámara de comercio de Duitama", "Nit: 123456789", "contacto@gestorccd.com", "Teléfono: (123) 456-7890"],
    ]
    tabla_encabezado = Table(encabezado_data, colWidths=[180, 180, 180, 180])
    estilo_encabezado = TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    tabla_encabezado.setStyle(estilo_encabezado)
    elements.append(tabla_encabezado)

    # Tabla usuario
    usuario = request.user
    data_usuario = [["Usuario", "Email", "Rol", "Cargo"]]
    data_usuario.append([
        usuario.username,
        usuario.email,
        getattr(usuario, 'role', 'No definido'),
        getattr(usuario, 'cargo', 'No definido'),
    ])
    table_usuario = Table(data_usuario, colWidths=[180, 180, 180, 180])
    style_usuario = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table_usuario.setStyle(style_usuario)
    elements.append(table_usuario)

    # Obtener artículos filtrados
    productos_filtrados = obtener_producto_bajo_stock_caf(request)

    if not productos_filtrados.exists():
        centered_style = ParagraphStyle(
            name="CenteredNormal",
            parent=styles["Normal"],
            alignment=TA_CENTER
        )
        no_results = Paragraph("No se encontraron artículos con bajo stock.", centered_style)
        elements.append(Spacer(1, 20))
        elements.append(no_results)
    else:
        # Tabla artículos
        data_productos = [["ID", "Nombre", "Marca", "Precio", "Cantidad", "Unidad de medida", "Proveedor", "Presentación"]]
        
        for producto in productos_filtrados:
            data_productos.append([
                wrap_text(str(producto.id)),
                wrap_text(producto.nombre),
                wrap_text(producto.marca),
                wrap_text("{:,}".format(producto.precio)),
                wrap_text(str(producto.cantidad)),
                wrap_text(str(producto.unidad_medida)),
                wrap_text(producto.proveedor),
                wrap_text(producto.presentacion),
            ])

        tabla_articulos = Table(data_productos, colWidths=[20, 140, 100, 110, 80, 70, 100])
        style_articulos = TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        tabla_articulos.setStyle(style_articulos)
        elements.append(tabla_articulos)

    # Marca de agua
    doc.build(elements, onFirstPage=draw_table_on_canvas, onLaterPages=draw_table_on_canvas)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Listado bajo stock CCD.pdf"'
    return response


def reporte_producto_bajo_stock_excel(request):
    # Parámetros de búsqueda
    q = request.GET.get('q', '').strip()
    fecha_inicio = request.GET.get('fecha_inicio', '').strip()
    fecha_fin = request.GET.get('fecha_fin', '').strip()
    marca = request.GET.get('marca', '').strip()

    # Queryset base
    productos = obtener_producto_bajo_stock_caf(request)

    # Aplicar filtros
    if q:
        productos = productos.filter(nombre__icontains=q)
    if fecha_inicio:
        try:
            productos = productos.filter(fecha_registro__gte=datetime.strptime(fecha_inicio, '%Y-%m-%d'))
        except ValueError:
            pass
    if fecha_fin:
        try:
            productos = productos.filter(fecha_registro__lte=datetime.strptime(fecha_fin, '%Y-%m-%d'))
        except ValueError:
            pass
    if marca:
        productos = productos.filter(marca__icontains=marca)

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Bajo stock de cafeteria CCD"

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 23
    ws.column_dimensions['H'].width = 23

    ws.row_dimensions[1].height = 60
    ws.row_dimensions[2].height = 30

    # Título
    ws.merge_cells('A1:H1')
    ws['A1'] = "GESTOR CCD"
    ws['A1'].font = Font(size=24, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Subtítulo
    ws.merge_cells('A2:H2')
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    ws['A2'] = f"Listado de cafeteria {fecha_actual}"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    # Bordes A1:H2
    for row in [1, 2]:
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f"{col}{row}"].border = border
            ws[f"{col}{row}"].alignment = Alignment(horizontal='center', vertical='center')

    # Encabezados
    headers = ["ID", "Nombre", "Marca", "Precio", "Cantidad", "Unidad de medida", "Proveedor", "Presentación"]
    ws.append(headers)
    header_fill = PatternFill(start_color="FF0056B3", end_color="FF0056B3", fill_type="solid")

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        cell = ws[f"{col}3"]
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    ws.auto_filter.ref = "A3:H3"

    # Agregar datos
    if not productos.exists():
        ws.merge_cells('A4:H4')
        cell = ws['A4']
        cell.value = "No se encontraron productos con los filtros aplicados."
        cell.font = Font(italic=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    else:
        for producto in productos:
            ws.append([
                str(producto.id),
                producto.nombre,
                producto.marca,
                "${:,}".format(producto.precio),
                str(producto.cantidad),
                str(producto.unidad_medida),
                producto.proveedor,
                producto.presentacion,
            ])
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=8):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Exportar
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Listado de bajo stock.xlsx"'
    wb.save(response)
    return response


@login_required
def crear_devolucion_caf(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)

    if request.method == 'POST':
        form = DevolucionFormCaf(request.POST, pedido_id=pedido_id)
        if form.is_valid():
            devolucion = form.save(commit=False)
            pedido_producto = devolucion.pedido_producto
            producto = pedido_producto.producto
            cantidad_devuelta = devolucion.cantidad_devuelta

            # Total ya devuelto
            total_devuelto = DevolucionCaf.objects.filter(
                pedido_producto__pedido=pedido,
                pedido_producto__producto=producto
            ).aggregate(total=Sum('cantidad_devuelta'))['total'] or 0

            if total_devuelto >= pedido_producto.cantidad:
                messages.error(request, "Ya has devuelto este producto en su totalidad.")
            elif total_devuelto + cantidad_devuelta > pedido_producto.cantidad:
                restante = pedido_producto.cantidad - total_devuelto
                messages.error(request, f"Solo puedes devolver hasta {restante} unidades.")
            else:
                devolucion.devuelto_por = request.user
                devolucion.save()

                # ✅ AUMENTAR STOCK EN INVENTARIO
                producto.cantidad += cantidad_devuelta
                producto.save()

                messages.success(request, "La devolución fue registrada exitosamente.")
                return redirect('cafeteria:mis_pedidos')
        else:
            messages.error(request, "Verifica que los campos sean válidos.")
    else:
        form = DevolucionFormCaf(pedido_id=pedido_id)

    return render(request, 'pedidos/devolver_producto.html', {
        'form': form,
        'pedido': pedido
    })