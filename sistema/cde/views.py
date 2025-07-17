from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib import messages
from django.urls import reverse
from django.db.models import Sum
from django.utils.timezone import localtime
from django.utils.html import escape
from django.conf import settings
from django.template.loader import render_to_string
from django.contrib.staticfiles import finders
from django.db.models import Q
from django.db import transaction
from django.utils import timezone
from datetime import datetime, timedelta
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO
from datetime import datetime
from django.http import HttpResponse
from django.db.models import Q
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from django.core.mail import send_mail
from cde.forms import LoginForm
from libreria.forms import CustomUserForm
from django.core.mail import EmailMultiAlternatives
from libreria.models import CustomUser
from datetime import datetime

from cafeteria.models import Productos
from django.conf import settings
from django.core.mail import send_mail
from cde.models import PedidoCde, PedidoProductoCde
from cde.forms import PedidoProductoCdeForm, LoginForm
from django.db.models import Q
from django.core.paginator import Paginator
from cafeteria.models import Productos

from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect
# Create your views here.
# Ejemplo para el login de papelería
def login_cde(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        user = authenticate(request, email=email, password=password)

        if user is not None:
            if user.is_active and getattr(user, 'acceso_cde', False):
                login(request, user)
                return redirect('cde:index_cde')
            else:
                messages.error(request, "No tienes permiso para acceder a este módulo.")
        else:
            messages.error(request, "Credenciales inválidas.")

    return render(request, 'login_cde/login_cde.html')

def logout_cde(request):
    logout(request)
    messages.success(request, "Has cerrado sesión correctamente.")
    return redirect(reverse('libreria:inicio'))
User = get_user_model()


def crear_pedido_cde(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': reverse('cde:index_cde')},
        {'name': 'Crear pedido CDE', 'url': reverse('cde:crear_pedido_cde')},
    ]

    if request.method == 'POST':
        try:
            estado = 'Confirmado' if request.user.role == 'Administrador' else 'Pendiente'

            pedido_cde = PedidoCde.objects.create(
                registrado_por=request.user,
                estado=estado,
                fecha_estado=timezone.now() if estado == 'Confirmado' else None
            )

            productos_ids = request.POST.getlist('producto')
            cantidades = request.POST.getlist('cantidad')
            eventos = request.POST.getlist('evento')

            area_usuario = getattr(request.user, 'area', 'No establecido')

            for producto_id, cantidad, evento in zip(productos_ids, cantidades, eventos):
                try:
                    if not producto_id or not cantidad or not evento:
                        continue

                    producto_id = int(producto_id)
                    cantidad = int(cantidad)

                    if cantidad <= 0:
                        raise ValueError("Cantidad debe ser mayor a cero")

                    producto = Productos.objects.get(id=producto_id)

                    if estado == 'Confirmado' and producto.cantidad < cantidad:
                        pedido_cde.delete()
                        messages.error(request, f"No hay suficiente stock para el producto: {producto.nombre} (disponible: {producto.cantidad}, solicitado: {cantidad})")
                        return redirect('cde:crear_pedido_cde')

                    PedidoProductoCde.objects.create(
                        pedido=pedido_cde,
                        producto=producto,
                        cantidad=cantidad,
                        evento=evento,
                        area=area_usuario,
                    )

                    if estado == 'Confirmado':
                        producto.cantidad -= cantidad
                        producto.save()

                except (ValueError, Productos.DoesNotExist) as e:
                    print(f"Error al procesar producto: {e}")
                    continue

            # Notificar a administradores
            admin_users = CustomUser.objects.filter(role='Administrador', is_active=True)
            admin_emails = [admin.email for admin in admin_users if admin.email]

            if admin_emails:
                admin_url = request.build_absolute_uri(reverse('cde:mis_pedidos_cde'))
                context = {
                    'usuario': request.user,
                    'pedido': pedido_cde,
                    'admin_url': admin_url,
                    'company_name': 'Gestor CCD',
                    'productos': PedidoProductoCde.objects.filter(pedido=pedido_cde),
                    'fecha_pedido': localtime(pedido_cde.fecha_pedido).strftime('%d/%m/%Y %H:%M') if pedido_cde.fecha_pedido else '',
                }

                html_message = render_to_string('pedidos_cde/email_notificacion_pedido.html', context)

                text_message = f"""
Hola Administrador,

Se ha registrado un nuevo pedido en el sistema del Centro de Eventos:

Usuario: {request.user.username}
Rol: {request.user.get_role_display()}
Área: {getattr(request.user, 'area', 'No especificada')}
ID del Pedido: {pedido_cde.id}
Estado: {estado}
Fecha: {pedido_cde.fecha_pedido.strftime('%d/%m/%Y %H:%M') if pedido_cde.fecha_pedido else 'N/D'}

Detalle de productos:
{chr(10).join([f"- {pp.producto.nombre} x {pp.cantidad} (Evento: {pp.evento})" for pp in PedidoProductoCde.objects.filter(pedido=pedido_cde)])}

Revisar el pedido: {admin_url}

Este es un mensaje automático, por favor no respondas.
"""

                try:
                    subject = f"Nuevo pedido {'confirmado' if estado == 'Confirmado' else 'pendiente'} - ID: {pedido_cde.id}"
                    msg = EmailMultiAlternatives(
                        subject,
                        text_message,
                        settings.DEFAULT_FROM_EMAIL,
                        admin_emails
                    )
                    msg.attach_alternative(html_message, "text/html")
                    msg.send()
                except Exception as e:
                    print(f"Error enviando correo: {e}")
                    messages.warning(request, "El pedido fue creado, pero no se pudo enviar el correo de notificación.")

            messages.success(request, f"El pedido fue registrado correctamente con estado '{estado}'.")
            return redirect('cde:mis_pedidos_cde')

        except Exception as e:
            print(f"Error al crear pedido: {e}")
            messages.error(request, "Ocurrió un error al procesar tu pedido. Por favor intenta nuevamente.")
            return render(request, 'pedidos_cde/pedidos_cde.html', {
                'productos': Productos.objects.all(),
                'breadcrumbs': breadcrumbs
            })

    productos = Productos.objects.all()
    return render(request, 'pedidos_cde/pedidos_cde.html', {
        'productos': productos,
        'breadcrumbs': breadcrumbs
    })
def ver_usuario_cde(request, id):
    breadcrumbs = [
        {'name': 'Inicio CDE', 'url': '/index_cde'},
        {'name': 'Ver usuario CDE', 'url': reverse('cde:ver_usuario_cde', kwargs={'id': id})},
    ]
    usuario = get_object_or_404(CustomUser, id=id)
    return render(request, 'usuario_cde/ver_perfil_cde.html', {'usuario': usuario, 'breadcrumbs': breadcrumbs})


def index_cde(request):
    breadcrumbs = [
        {'name': 'Inicio CDE', 'url': '/index_cde'},
    ]
    return render(request, 'index_cde/index_cde.html',{'breadcrumbs': breadcrumbs})


def mis_pedidos_cde(request):
    breadcrumbs = [
        {'name': 'Inicio CDE', 'url': '/index_cde'},
        {'name': 'Mis pedidos cde', 'url': reverse('cde:mis_pedidos_cde')},
    ]

    query = request.GET.get('q', '').strip()
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    pedidos = PedidoCde.objects.filter(registrado_por=request.user).order_by('-fecha_pedido')

    if query:
        pedidos = pedidos.filter(
            Q(id__icontains=query) |
            Q(estado__icontains=query) |
            Q(registrado_por__username__icontains=query) |
            Q(productos__producto__nombre__icontains=query) |
            Q(productos__area__icontains=query) |
            Q(productos__evento__icontains=query) |
            Q(productos__cantidad__icontains=query)
        ).distinct()

    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
            pedidos = pedidos.filter(fecha_pedido__gte=fecha_inicio)
        except ValueError:
            pedidos = PedidoCde.objects.none()

    if fecha_fin_str:
        try:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d') + timedelta(days=1)
            pedidos = pedidos.filter(fecha_pedido__lt=fecha_fin)
        except ValueError:
            pedidos = PedidoCde.objects.none()

    paginator = Paginator(pedidos, 4)
    page_number = request.GET.get('page')
    pedidos_page = paginator.get_page(page_number)

    # Crear un string con los parámetros de búsqueda para la paginación
    query_params = ''
    if query:
        query_params += f'&q={query}'
    if fecha_inicio_str:
        query_params += f'&fecha_inicio={fecha_inicio_str}'
    if fecha_fin_str:
        query_params += f'&fecha_fin={fecha_fin_str}'

    return render(request, 'pedidos_cde/mis_pedidos_cde.html', {
        'pedidos': pedidos_page,
        'breadcrumbs': breadcrumbs,
        'query_params': query_params,
        'current_query': query,
        'current_fecha_inicio': fecha_inicio_str,
        'current_fecha_fin': fecha_fin_str,
    })

@csrf_exempt  # Solo si no puedes usar {% csrf_token %} en el formulario
@require_POST
def cambiar_estado_pedido_cde(request, pedido_id):
    pedido = get_object_or_404(PedidoCde, id=pedido_id)
    nuevo_estado = request.POST.get('estado')

    if not nuevo_estado:
        messages.error(request, 'No se especificó un nuevo estado.')
        return redirect('cde:pedidos_pendientes_cde')

    if nuevo_estado == 'Confirmado':
        try:
            with transaction.atomic():
                productos_pedido = pedido.productos.select_related('producto').all()
                productos_sin_stock = []

                for item in productos_pedido:
                    producto = item.producto
                    if producto.cantidad < item.cantidad:
                        productos_sin_stock.append(
                            f"{producto.nombre} (Solicitados: {item.cantidad}, Disponibles: {producto.cantidad})"
                        )

                if productos_sin_stock:
                    pedido.estado = 'Cancelado'
                    pedido.fecha_estado = timezone.now()
                    pedido.save()
                    messages.error(
                        request,
                        f"Pedido cancelado. Stock insuficiente: {', '.join(productos_sin_stock)}"
                    )
                    return redirect('cde:pedidos_pendientes_cde')

                for item in productos_pedido:
                    producto = item.producto
                    producto.cantidad -= item.cantidad
                    producto.save()

                pedido.estado = 'Confirmado'
                pedido.fecha_estado = timezone.now()
                pedido.save()
                messages.success(request, f'Pedido #{pedido.id} confirmado correctamente.')

        except Exception as e:
            messages.error(request, f'Ocurrió un error al confirmar el pedido: {str(e)}')
            return redirect('cde:pedidos_pendientes_cde')

    else:
        pedido.estado = nuevo_estado
        pedido.fecha_estado = timezone.now()
        pedido.save()
        messages.success(request, f'Estado del pedido #{pedido.id} actualizado a "{nuevo_estado}".')

    if pedido.registrado_por and pedido.registrado_por.email:
        productos_lista = "\n".join(
            f"<li class='articulo-item'>{escape(item.cantidad)} {escape(item.producto.nombre)}</li>"
            for item in pedido.productos.all()
        )

        html_content = f"""
        <!DOCTYPE html>
        <html lang="es">
        <head>
            <meta charset="UTF-8">
            <title>Estado de Pedido Actualizado</title>
            <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            line-height: 1.6;
            color: #333333;
            max-width: 600px;
            margin: 0 auto;
            padding: 20px;
            background-color: #f5f5f5;
        }}
        .email-container {{
            background-color: #ffffff;
            border-radius: 8px;
            box-shadow: 0 2px 10px rgba(0, 0, 0, 0.1);
            overflow: hidden;
        }}
        .header {{
            color: white;
            padding: 20px;
            text-align: center;
        }}
        .logo {{
            max-width: 180px;
            height: auto;
            margin-bottom: 15px;
        }}
        .content {{
            padding: 25px;
        }}
        .status {{
            font-size: 18px;
            font-weight: bold;
            margin: 15px 0;
            padding: 10px;
            border-radius: 5px;
            text-align: center;
        }}
        .status.Confirmado {{
            color: #155724;
        }}
        .status.Cancelado {{
            background-color: #f8d7da;
            color: #721c24;
        }}
        .status.Pendiente {{
            color: #856404;
        }}
        .info-box {{
            background-color: #f8f9fa;
            border-left: 4px solid #007bff;
            padding: 15px;
            margin: 20px 0;
        }}
        .info-label {{
            font-weight: bold;
            color: #495057;
            display: inline-block;
            min-width: 150px;
        }}
        .articulos-list {{
            margin: 15px 0;
            padding-left: 20px;
        }}
        .articulo-item {{
            margin-bottom: 8px;
        }}
        .footer {{
            text-align: center;
            padding: 15px;
            font-size: 12px;
            color: #6c757d;
            border-top: 1px solid #e9ecef;
            background-color: #f8f9fa;
        }}
    </style>
        </head>
        <body>
            <div class="email-container">
                <div class="header">
                    <img src="https://ccduitama.org.co/wp-content/uploads/2021/05/LOGOCCD-TRANSPARENCIA.png" alt="Logo CCD" class="logo">
                    <h2>Actualización de Estado de Pedido</h2>
                </div>
                
                <div class="content">
                    <p>Hola <strong>{escape(pedido.registrado_por.username)}</strong>,</p>
                    <p>Te informamos que el estado de tu pedido en el módulo de centro de eventos ha sido actualizado:</p>
                    
                    <div class="info-box">
                        <p><span class="info-label">Número de Pedido:</span> #{pedido.id}</p>
                        <p><span class="info-label">Fecha de Actualización:</span> {pedido.fecha_estado.strftime('%d/%m/%Y %H:%M')}</p>
                        <p><span class="info-label">Nuevo Estado:</span> <span class="status {pedido.estado}">{pedido.estado.upper()}</span></p>

                        <h3>Detalle del Pedido:</h3>
                        <ul class="articulos-list">
                            {productos_lista}
                        </ul>
                    </div>
                    
                    <p>Para más información, puedes acceder al sistema de gestión de pedidos.</p>
                </div>
                
                <div class="footer">
                    <p>Este es un mensaje automático, por favor no respondas a este correo.</p>
                    <p>© {timezone.now().year} Gestor Cafetería - Todos los derechos reservados</p>
                </div>
            </div>
        </body>
        </html>
        """

        text_content = f"""
        Actualización de Estado de Pedido

        Hola {pedido.registrado_por.username},

        Tu pedido #{pedido.id} ha sido actualizado a {pedido.estado.upper()}.

        Fecha de actualización: {pedido.fecha_estado.strftime('%d/%m/%Y %H:%M')}

        Detalle del Pedido:
        {"".join([f"- {item.cantidad} x {item.producto.nombre}\n" for item in pedido.productos.all()])}

        Gracias por usar nuestro sistema.
        © {timezone.now().year} Gestor Cafetería
        """

        try:
            send_mail(
                subject=f'Pedido #{pedido.id} - Estado actualizado a {pedido.estado.upper()}',
                message=text_content,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[pedido.registrado_por.email],
                html_message=html_content
            )
        except Exception as e:
            print(f"Error enviando email: {str(e)}")

    return redirect('cde:pedidos_pendientes_cde')

def pedidos_pendientes_cde(request):
    breadcrumbs = [
        {'name': 'Inicio CDE', 'url': '/index_cde'},
        {'name': 'Pedidos pendientes cde', 'url': reverse('cde:pedidos_pendientes_cde')},
    ]
    
    query = request.GET.get('q', '').strip()
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    # Filtrar solo pedidos pendientes
    pedidos = PedidoCde.objects.filter(estado='Pendiente').order_by('-fecha_pedido')

    if query:
        pedidos = pedidos.filter(
            Q(id__icontains=query) |
            Q(estado__icontains=query) |
            Q(registrado_por__username__icontains=query) |
            Q(productos__producto__nombre__icontains=query) |
            Q(productos__area__icontains=query) |
            Q(productos__evento__icontains=query) |
            Q(productos__cantidad__icontains=query)
        ).distinct()

    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            pedidos = pedidos.filter(fecha_pedido__gte=fecha_inicio)
        except ValueError:
            pass  # No hacer nada si la fecha es inválida

    if fecha_fin_str:
        try:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date() + timedelta(days=1)
            pedidos = pedidos.filter(fecha_pedido__lt=fecha_fin)
        except ValueError:
            pass  # No hacer nada si la fecha es inválida

    paginator = Paginator(pedidos, 4)
    page_number = request.GET.get('page')
    pedidos_page = paginator.get_page(page_number)

    # Crear string con parámetros de búsqueda para la paginación
    query_params = ''
    if query:
        query_params += f'&q={query}'
    if fecha_inicio_str:
        query_params += f'&fecha_inicio={fecha_inicio_str}'
    if fecha_fin_str:
        # Mostrar la fecha original en los parámetros (sin el timedelta)
        query_params += f'&fecha_fin={fecha_fin_str}'

    return render(request, 'pedidos_cde/confirmar_pedido_cde.html', {
        'pedidos': pedidos_page,
        'breadcrumbs': breadcrumbs,
        'query_params': query_params,
        'current_query': query,
        'current_fecha_inicio': fecha_inicio_str,
        'current_fecha_fin': fecha_fin_str[:-10] if fecha_fin_str else None,  # Remover el timedelta para mostrar
    })


def listado_pedidos_cde(request):
    breadcrumbs = [
        {'name': 'Inicio CDE', 'url': '/index_cde'},
        {'name': 'Listado de pedidos cde', 'url': reverse('cde:lista_pedidos_cde')},
    ]

    query = request.GET.get('q', '').strip()
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    pedidos = PedidoCde.objects.filter(estado__in=['Confirmado', 'Cancelado']).order_by('-fecha_pedido')

    if query:
        pedidos = pedidos.filter(
            Q(id__icontains=query) |
            Q(estado__icontains=query) |
            Q(fecha_pedido__icontains=query) |
            Q(fecha_estado__icontains=query) |
            Q(registrado_por__username__icontains=query) |
            Q(productos__producto__nombre__icontains=query) |
            Q(productos__cantidad__icontains=query) |
            Q(productos__area__icontains=query) |
            Q(productos__evento__icontains=query)
        ).distinct()

    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            pedidos = pedidos.filter(fecha_pedido__gte=fecha_inicio)
        except ValueError:
            pedidos = PedidoCde.objects.none()

    if fecha_fin_str:
        try:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
            pedidos = pedidos.filter(fecha_pedido__lt=fecha_fin + timedelta(days=1))
        except ValueError:
            pedidos = PedidoCde.objects.none()

    paginator = Paginator(pedidos, 4)
    page_number = request.GET.get('page')
    pedidos_page = paginator.get_page(page_number)

    # Crear string con parámetros de búsqueda para la paginación
    query_params = ''
    if query:
        query_params += f'&q={query}'
    if fecha_inicio_str:
        query_params += f'&fecha_inicio={fecha_inicio_str}'
    if fecha_fin_str:
        query_params += f'&fecha_fin={fecha_fin_str}'

    return render(request, 'pedidos_cde/lista_pedidos_cde.html', {
        'pedidos': pedidos_page,
        'breadcrumbs': breadcrumbs,
        'query_params': query_params,
        'current_query': query,
        'current_fecha_inicio': fecha_inicio_str,
        'current_fecha_fin': fecha_fin_str,
    })

#pdf
def draw_table_on_canvas(canvas, doc):
    # Marca de agua
    watermark_path = finders.find('imagen/LOGO.png')
    if watermark_path:
        canvas.saveState()
        canvas.setFillColor(colors.Color(1, 1, 1, alpha=0.3))
        canvas.setStrokeColor(colors.Color(1, 1, 1, alpha=0.3))
        canvas.drawImage(watermark_path,
                         x=(doc.pagesize[0] - 600) / 2,
                         y=(doc.pagesize[1] - 600) / 2,
                         width=600, height=600, mask='auto')
        canvas.restoreState()
def wrap_text_p(text, max_len=26
                ):
    parts = [text[i:i+max_len] for i in range(0, len(text), max_len)]
    for i in range(len(parts) - 1):
        parts[i] += '-'  # Agrega guion al final de todas menos la última
    return '\n'.join(parts)
def get_pedidos_filtrados_cde(request):
    query = request.GET.get('q')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    pedidos = PedidoCde.objects.filter(estado__in=['Confirmado', 'Cancelado']).order_by('-fecha_pedido')

    if query:
        pedidos = pedidos.filter(
            Q(id__icontains=query) |
            Q(estado__icontains=query) |
            Q(fecha_pedido__icontains=query) |
            Q(fecha_estado__icontains=query) |
            Q(registrado_por__username__icontains=query) |
            Q(productos__producto__nombre__icontains=query) |
            Q(productos__cantidad__icontains=query) |
            Q(productos__area__icontains=query) |
            Q(productos__evento__icontains=query)
        ).distinct()

    if fecha_inicio and fecha_fin:
        pedidos = pedidos.filter(fecha_pedido__range=[fecha_inicio, fecha_fin])

    return pedidos
def reporte_pedidos_pdf_cde(request):

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=40,
        rightMargin=40,
        topMargin=40,
        bottomMargin=40
    )

    doc.title = "Listado de pedidos  centro de eventos"
    doc.author = "CCD"
    doc.subject = "Listado de pedidos cde"
    doc.creator = "Sistema de Gestión CCD"

    elements = []
    styles = getSampleStyleSheet()

    # Título principal
    titulo = Paragraph("REPORTE DE PEDIDOS DE CENTRO DE EVENTOS", styles["Title"])
    elements.append(titulo)

    # Encabezado institucional
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    encabezado_data = [
        ["GESTOR CCD", "Lista de usuarios", "Correo: gestorccd@gmail.com", f"Fecha: {fecha_actual}"],
        ["Cámara de comercio de Duitama", "Nit: 123456789", "(Correo de la camara)", "Teléfono: (tel. camara)"],
    ]
    tabla_encabezado = Table(encabezado_data, colWidths=[180, 180, 180, 180])
    estilo_encabezado = TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    tabla_encabezado.setStyle(estilo_encabezado)
    elements.append(tabla_encabezado)

    # Datos del usuario
    usuario = request.user
    data_usuario = [["Usuario:", "Email:", "Rol:", "Cargo:"]]
    data_usuario.append([
        usuario.username,
        usuario.email,
        getattr(usuario, 'role', 'No definido'),
        getattr(usuario, 'cargo', 'No definido'),
    ])
    table_usuario = Table(data_usuario, colWidths=[180, 180, 180, 180])
    style_usuario = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table_usuario.setStyle(style_usuario)
    elements.append(table_usuario)

    # Obtener filtros
    q = request.GET.get('q', '')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    # Fetch pedidos with prefetch for productos and producto
    pedidos = PedidoCde.objects.prefetch_related('productos__producto') \
    .filter(estado__in=['Confirmado', 'Cancelado'])


    if q:
        pedidos = pedidos.filter(
            Q(id__icontains=q) |
            Q(estado__icontains=q) |
            Q(fecha_pedido__icontains=q) |
            Q(fecha_estado__icontains=q) |
            Q(registrado_por__username__icontains=q) |
            Q(productos__producto__nombre__icontains=q) |
            Q(productos__cantidad__icontains=q) |
            Q(productos__area__icontains=q) |
            Q(productos__evento__icontains=q)
        ).distinct()

    if fecha_inicio:
        try:
            pedidos = pedidos.filter(fecha_pedido__gte=datetime.strptime(fecha_inicio, '%Y-%m-%d'))
        except ValueError:
            pass  # Handle invalid date format gracefully
    if fecha_fin:
        try:
            pedidos = pedidos.filter(fecha_pedido__lte=datetime.strptime(fecha_fin, '%Y-%m-%d'))
        except ValueError:
            pass  # Handle invalid date format gracefully

    if not pedidos.exists():
        centered_style = ParagraphStyle(
            name="CenteredNormal",
            parent=styles["Normal"],
            alignment=TA_CENTER,
        )
        no_results = Paragraph("No se encontraron pedidos.", centered_style)
        elements.append(no_results)
    else:
        # Encabezado de la tabla
        data_pedidos = [["ID Pedido", "Fecha", "Estado", "Registrado Por", "Productos", "Área"]]

        # Filas de pedidos
        for pedido in pedidos:
            try:
                productos = pedido.productos.all()
                if productos.exists():
                    productos_raw = ", ".join([
                        f"{pa.producto.nombre} x {pa.cantidad}"
                        for pa in productos
                    ])
                    areas = set(pa.area for pa in productos if pa.area and pa.area != 'No establecido')
                    area_raw = ", ".join(areas) if areas else 'No establecido'
                else:
                    productos_raw = 'Sin productos'
                    area_raw = 'Sin área'

                data_pedidos.append([
                    wrap_text_p(str(pedido.id)),
                    wrap_text_p(pedido.fecha_pedido.strftime('%d-%m-%Y')),
                    wrap_text_p(pedido.get_estado_display()),
                    wrap_text_p(pedido.registrado_por.username if pedido.registrado_por else 'No definido'),
                    wrap_text_p(productos_raw),
                    wrap_text_p(area_raw)
                ])
            except Exception as e:
                print(f"Error processing pedido {pedido.id}: {str(e)}")
                data_pedidos.append([
                    wrap_text_p(str(pedido.id)),
                    wrap_text_p(pedido.fecha_pedido.strftime('%d-%m-%Y')),
                    wrap_text_p(pedido.get_estado_display()),
                    wrap_text_p(pedido.registrado_por.username if pedido.registrado_por else 'No definido'),
                    wrap_text_p('Error al cargar productos'),
                    wrap_text_p('Error al cargar área')
                ])

        # Crear la tabla de pedidos
        tabla_productos = Table(data_pedidos, colWidths=[60, 100, 100, 160, 200, 100])
        style_productos = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        tabla_productos.setStyle(style_productos)
        elements.append(tabla_productos)

    # Construir el PDF
    doc.build(elements, onFirstPage=draw_table_on_canvas, onLaterPages=draw_table_on_canvas)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Lista de pedidos Gestor CCD.pdf"'
    return response

def reporte_pedidos_excel_cde(request):
    pedidos = get_pedidos_filtrados_cde(request).prefetch_related('productos__producto')

    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos CCD"

    columnas_anchos = {
        'A': 10,
        'B': 15,
        'C': 15,
        'D': 25,
        'E': 50,
        'F': 30
    }
    for col, width in columnas_anchos.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 50
    ws.row_dimensions[2].height = 30

    ws.merge_cells('A1:F1')
    ws['A1'] = "GESTOR CCD"
    ws['A1'].font = Font(size=24, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:F2')
    ws['A2'] = "Listado de Pedidos"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=6):
        for cell in row:
            cell.border = thin_border

    headers = ['ID', 'Fecha', 'Estado', 'Registrado Por', 'Productos', 'Áreas']
    ws.append(headers)

    header_fill = PatternFill(start_color="FF0056B3", end_color="FF0056B3", fill_type="solid")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.auto_filter.ref = "A3:F3"

    # Agregar datos o mensaje si no hay pedidos
    if not pedidos.exists():
        ws.merge_cells('A4:F4')
        cell = ws['A4']
        cell.value = "No se encontraron pedidos."
        cell.font = Font(italic=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws.row_dimensions[4].height = 40
    else:
        for pedido in pedidos:
            productos_raw = ", ".join([
                f"{pa.cantidad} {pa.producto.nombre} para {pa.evento}"
                for pa in pedido.productos.all()
            ]) or 'Sin artículos'

            areas_raw = ", ".join(set([
                str(pa.area) for pa in pedido.productos.all()
            ])) or 'Sin área'

            usuario = pedido.registrado_por.username if pedido.registrado_por else 'No definido'

            ws.append([
                pedido.id,
                pedido.fecha_pedido.strftime('%Y-%m-%d'),
                pedido.get_estado_display(),
                usuario,
                productos_raw,
                areas_raw,
            ])

        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=6):
            for i, cell in enumerate(row, start=1):
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=True if i in [4, 5, 6] else False
                )

        for i in range(4, ws.max_row + 1):
            ws.row_dimensions[i].height = 60

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Reporte pedidos cde.xlsx"'
    wb.save(response)
    return response

def get_pedidos_filtrados_pendientes_cde(request):
    query = request.GET.get('q')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    pedidos = PedidoCde.objects.filter(estado='Pendiente').order_by('-fecha_pedido')

    if query:
        pedidos = pedidos.filter(
            Q(id__icontains=query) |
            Q(estado__icontains=query) |
            Q(fecha_pedido__icontains=query) |
            Q(fecha_estado__icontains=query) |
            Q(registrado_por__username__icontains=query) |
            Q(productos__producto__nombre__icontains=query) |
            Q(productos__cantidad__icontains=query) |
            Q(productos__area__icontains=query) |
            Q(productos__evento__icontains=query)
        ).distinct()

    if fecha_inicio and fecha_fin:
        pedidos = pedidos.filter(fecha_pedido__range=[fecha_inicio, fecha_fin])

    return pedidos

def reporte_pedidos_pendientes_pdf_cde(request):

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=40,
        rightMargin=40,
        topMargin=40,
        bottomMargin=40
    )

    doc.title = "Listado de pedidos pendientes centro de eventos"
    doc.author = "CCD"
    doc.subject = "Listado de pedidos cde"
    doc.creator = "Sistema de Gestión CCD"

    elements = []
    styles = getSampleStyleSheet()

    # Título principal
    titulo = Paragraph("REPORTE DE PEDIDOS PENDIENTES DE CENTRO DE EVENTOS", styles["Title"])
    elements.append(titulo)

    # Encabezado institucional
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    encabezado_data = [
        ["GESTOR CCD", "Lista de usuarios", "Correo: gestorccd@gmail.com", f"Fecha: {fecha_actual}"],
        ["Cámara de comercio de Duitama", "Nit: 123456789", "(Correo de la camara)", "Teléfono: (tel. camara)"],
    ]
    tabla_encabezado = Table(encabezado_data, colWidths=[180, 180, 180, 180])
    estilo_encabezado = TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    tabla_encabezado.setStyle(estilo_encabezado)
    elements.append(tabla_encabezado)

    # Datos del usuario
    usuario = request.user
    data_usuario = [["Usuario:", "Email:", "Rol:", "Cargo:"]]
    data_usuario.append([
        usuario.username,
        usuario.email,
        getattr(usuario, 'role', 'No definido'),
        getattr(usuario, 'cargo', 'No definido'),
    ])
    table_usuario = Table(data_usuario, colWidths=[180, 180, 180, 180])
    style_usuario = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table_usuario.setStyle(style_usuario)
    elements.append(table_usuario)

    # Obtener filtros
    q = request.GET.get('q', '')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    # Fetch pedidos with prefetch for productos and producto
    pedidos = get_pedidos_filtrados_pendientes_cde(request).prefetch_related('productos__producto')

    if q:
        pedidos = pedidos.filter(
            Q(id__icontains=q) |
            Q(estado__icontains=q) |
            Q(fecha_pedido__icontains=q) |
            Q(fecha_estado__icontains=q) |
            Q(registrado_por__username__icontains=q) |
            Q(productos__producto__nombre__icontains=q) |
            Q(productos__cantidad__icontains=q) |
            Q(productos__area__icontains=q) |
            Q(productos__evento__icontains=q)
        ).distinct()

    if fecha_inicio:
        try:
            pedidos = pedidos.filter(fecha_pedido__gte=datetime.strptime(fecha_inicio, '%Y-%m-%d'))
        except ValueError:
            pass  # Handle invalid date format gracefully
    if fecha_fin:
        try:
            pedidos = pedidos.filter(fecha_pedido__lte=datetime.strptime(fecha_fin, '%Y-%m-%d'))
        except ValueError:
            pass  # Handle invalid date format gracefully

    if not pedidos.exists():
        centered_style = ParagraphStyle(
            name="CenteredNormal",
            parent=styles["Normal"],
            alignment=TA_CENTER,
        )
        no_results = Paragraph("No se encontraron pedidos pendientes.", centered_style)
        elements.append(no_results)
    else:
        # Encabezado de la tabla
        data_pedidos = [["ID Pedido", "Fecha", "Estado", "Registrado Por", "Productos", "Área"]]

        # Filas de pedidos
        for pedido in pedidos:
            try:
                productos = pedido.productos.all()
                if productos.exists():
                    productos_raw = ", ".join([
                        f"{pa.producto.nombre} x {pa.cantidad}"
                        for pa in productos
                    ])
                    areas = set(pa.area for pa in productos if pa.area and pa.area != 'No establecido')
                    area_raw = ", ".join(areas) if areas else 'No establecido'
                else:
                    productos_raw = 'Sin productos'
                    area_raw = 'Sin área'

                data_pedidos.append([
                    wrap_text_p(str(pedido.id)),
                    wrap_text_p(pedido.fecha_pedido.strftime('%d-%m-%Y')),
                    wrap_text_p(pedido.get_estado_display()),
                    wrap_text_p(pedido.registrado_por.username if pedido.registrado_por else 'No definido'),
                    wrap_text_p(productos_raw),
                    wrap_text_p(area_raw)
                ])
            except Exception as e:
                print(f"Error processing pedido {pedido.id}: {str(e)}")
                data_pedidos.append([
                    wrap_text_p(str(pedido.id)),
                    wrap_text_p(pedido.fecha_pedido.strftime('%d-%m-%Y')),
                    wrap_text_p(pedido.get_estado_display()),
                    wrap_text_p(pedido.registrado_por.username if pedido.registrado_por else 'No definido'),
                    wrap_text_p('Error al cargar productos'),
                    wrap_text_p('Error al cargar área')
                ])

        # Crear la tabla de pedidos
        tabla_productos = Table(data_pedidos, colWidths=[60, 100, 100, 160, 200, 100])
        style_productos = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        tabla_productos.setStyle(style_productos)
        elements.append(tabla_productos)

    # Construir el PDF
    doc.build(elements, onFirstPage=draw_table_on_canvas, onLaterPages=draw_table_on_canvas)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Lista de pedidos pendientes Gestor CCD.pdf"'
    return response

def reporte_pedidos_pendientes_excel_cde(request):
    pedidos = get_pedidos_filtrados_pendientes_cde(request).prefetch_related('productos__producto')

    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos CCD"

    columnas_anchos = {
        'A': 10,
        'B': 15,
        'C': 15,
        'D': 25,
        'E': 50,
        'F': 30
    }
    for col, width in columnas_anchos.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 50
    ws.row_dimensions[2].height = 30

    ws.merge_cells('A1:F1')
    ws['A1'] = "GESTOR CCD"
    ws['A1'].font = Font(size=24, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:F2')
    ws['A2'] = "Listado de Pedidos"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=6):
        for cell in row:
            cell.border = thin_border

    headers = ['ID', 'Fecha', 'Estado', 'Registrado Por', 'Productos', 'Áreas']
    ws.append(headers)

    header_fill = PatternFill(start_color="FF0056B3", end_color="FF0056B3", fill_type="solid")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.auto_filter.ref = "A3:F3"

    # Agregar datos o mensaje si no hay pedidos
    if not pedidos.exists():
        ws.merge_cells('A4:F4')
        cell = ws['A4']
        cell.value = "No se encontraron pedidos."
        cell.font = Font(italic=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws.row_dimensions[4].height = 40
    else:
        for pedido in pedidos:
            productos_raw = ", ".join([
                f"{pa.cantidad} {pa.producto.nombre} para {pa.evento}"
                for pa in pedido.productos.all()
            ]) or 'Sin artículos'

            areas_raw = ", ".join(set([
                str(pa.area) for pa in pedido.productos.all()
            ])) or 'Sin área'

            usuario = pedido.registrado_por.username if pedido.registrado_por else 'No definido'

            ws.append([
                pedido.id,
                pedido.fecha_pedido.strftime('%Y-%m-%d'),
                pedido.get_estado_display(),
                usuario,
                productos_raw,
                areas_raw,
            ])

        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=6):
            for i, cell in enumerate(row, start=1):
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=True if i in [4, 5, 6] else False
                )

        for i in range(4, ws.max_row + 1):
            ws.row_dimensions[i].height = 60

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Reporte pedidos pendientes cde.xlsx"'
    wb.save(response)
    return response

def index_estadistica_cde(request):
    breadcrumbs = [
     {'name': 'Inicio CDE', 'url': '/index_cde'},
]

    return render(request, 'estadisticas_cde/index_estadistica_cde.html', {'breadcrumbs': breadcrumbs})

def grafica_pedidos_cde(request):
    breadcrumbs = [
        {'name': 'Inicio CDE', 'url': '/index_cde'},
        {'name': 'Estadísticas', 'url': reverse('cde:index_estadistica_cde')},
        {'name': 'Gráfico de pedidos CDE', 'url': reverse('cde:grafica_pedidos_cde')},
    ]

    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    productos = PedidoProductoCde.objects.select_related('pedido', 'producto', 'pedido__registrado_por')

    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, "%Y-%m-%d")
            productos = productos.filter(pedido__fecha_pedido__gte=fecha_inicio)
        except ValueError:
            fecha_inicio = None  # puedes manejar errores si quieres
    else:
        fecha_inicio = None

    if fecha_fin_str:
        try:
            fecha_fin = datetime.strptime(fecha_fin_str, "%Y-%m-%d") + timedelta(days=1)
            productos = productos.filter(pedido__fecha_pedido__lt=fecha_fin)
        except ValueError:
            fecha_fin = None
    else:
        fecha_fin = None

    datos = productos.values(
        'pedido__registrado_por__username',
        'producto__nombre'
    ).annotate(
        total_cantidad=Sum('cantidad')
    ).order_by('pedido__registrado_por__username', 'producto__nombre')

    etiquetas = [item['producto__nombre'] for item in datos]
    cantidades = [item['total_cantidad'] for item in datos]

    return render(request, 'estadisticas_cde/grafico_pedidos_cde.html', {
        'nombres': etiquetas,
        'cantidades': cantidades,
        'breadcrumbs': breadcrumbs,
        'fecha_inicio': fecha_inicio_str,
        'fecha_fin': fecha_fin_str,
    })
def grafica_estado_pedido_cde(request):
    breadcrumbs = [
        {'name': 'Inicio CDE', 'url': '/index_cde'},
        {'name': 'Estadísticas', 'url': reverse('cde:index_estadistica_cde')},
        {'name': 'Gráfico de estado de pedidos CDE', 'url': reverse('cde:grafica_estado_pedido_cde')},
    ]

    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    pedidos = PedidoCde.objects.all()

    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, "%Y-%m-%d")
            pedidos = pedidos.filter(fecha_pedido__gte=fecha_inicio)
        except ValueError:
            fecha_inicio = None
    else:
        fecha_inicio = None

    if fecha_fin_str:
        try:
            fecha_fin = datetime.strptime(fecha_fin_str, "%Y-%m-%d") + timedelta(days=1)
            pedidos = pedidos.filter(fecha_pedido__lt=fecha_fin)
        except ValueError:
            fecha_fin = None
    else:
        fecha_fin = None

    pendientes = pedidos.filter(estado='Pendiente').count()
    confirmados = pedidos.filter(estado='Confirmado').count()
    cancelados = pedidos.filter(estado='Cancelado').count()

    nombres = ['Pendiente', 'Confirmado', 'Cancelado']
    cantidades = [pendientes, confirmados, cancelados]

    return render(request, 'estadisticas_cde/grafica_estado_pendiente_cde.html', {
        'nombres': nombres,
        'cantidades': cantidades,
        'breadcrumbs': breadcrumbs,
        'fecha_inicio': fecha_inicio_str,
        'fecha_fin': fecha_fin_str
    })