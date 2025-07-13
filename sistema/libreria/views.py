from django.shortcuts import render, redirect
from django.contrib.auth import update_session_auth_hash
from django.contrib import messages
from django.core.mail import send_mail
from django.db import IntegrityError
from django.core.exceptions import ValidationError
from django.conf import settings
from django.contrib.auth import authenticate, login
from libreria.forms import CustomUserForm, CustomUserEditForm, CustomPasswordChangeForm
from django.shortcuts import get_object_or_404
from .models import CustomUser
from django.contrib.auth import get_user_model
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.core.mail import send_mail
from django.shortcuts import render, redirect
from django.urls import reverse
from django.template.loader import render_to_string
from django.contrib.auth.forms import SetPasswordForm
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.db.models import Q
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import get_user_model
from django.http import HttpResponse
from openpyxl import Workbook
import os
from django.http import FileResponse, Http404
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
# Create your views here.
# Página de inicio
User = get_user_model()
def acceso_denegado(request):
    return render(request, 'acceso_denegado.html')
def error_404_view(request, exception):
    return render(request, 'acceso_denegado.html', status=404)
def timeouterror(request):
    try:
        # Simulación de una operación que puede causar un TimeoutError
        # Aquí va tu lógica real, como una conexión a red, base de datos externa, etc.
        raise TimeoutError("Error de tiempo de espera")  # Simulación

        # Si no ocurre error, puedes devolver otro template si lo deseas
        return render(request, 'exito.html')

    except TimeoutError:
        # Solo captura TimeoutError y redirige a lan_error.html
        return render(request, 'lan_error.html')
    
def inicio(request):
    return render(request, 'index/index.html')

def manual_usuario_view(request):
    pdf_path = os.path.join(settings.BASE_DIR, 'libreria', 'static', 'manual','manual_usuario_papeleria.pdf')
    try:
        return FileResponse(open(pdf_path, 'rb'),
    content_type='application/pdf')
    except FileNotFoundError:
        raise Http404("Manual no encontrado")
    
def crear_usuario(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Crear usuario', 'url': '/crear_usuario'},
    ]
    admin_exists = CustomUser.objects.exists()

    if request.method == 'POST':
        form = CustomUserForm(request.POST)

        if form.is_valid():
            try:
                user = form.save(commit=False)
                role = user.role

                # ✅ MODIFICADO: Permitir hasta 1 administrador activo en CREACIÓN
                if role == 'Administrador':
                    admin_count = CustomUser.objects.filter(
                        role='Administrador',
                        is_active=True
                    ).count()

                    limit = 1  # CAMBIADO: Límite de 1 para creación
                    if admin_count >= limit:
                        messages.error(
                            request,
                            f"Límite de administradores alcanzado. Solo se permite {limit} administrador activo durante el registro."
                        )
                        return redirect('libreria:crear_usuario')

                    # Asignar permisos a administradores
                    user.acceso_pap = True
                    user.acceso_caf = True
                    user.acceso_cde = True
                else:
                    # Restringir permisos para empleados
                    user.acceso_pap = False
                    user.acceso_caf = False
                    user.acceso_cde = False

                user.save()

                # Notificación a administradores
                admin_emails = CustomUser.objects.filter(
                    role='Administrador',
                    is_active=True
                ).exclude(pk=user.pk).values_list('email', flat=True)

                cargo = request.POST.get("cargo", "").strip()
                email_nuevo = user.email

                if admin_emails:
                    subject = "Nuevo usuario registrado en Gestor CCD"
                    message = (
                        f"Detalles del nuevo usuario:\n\n"
                        f"• Nombre de usuario: {user.username}\n"
                        f"• Rol: {role}\n"
                        f"• Cargo: {cargo}\n"
                        f"• Email: {email_nuevo}\n\n"
                        f"Permisos asignados:\n"
                        f"• Papelería: {'Sí' if user.acceso_pap else 'No'}\n"
                        f"• Cafetería: {'Sí' if user.acceso_caf else 'No'}\n"
                        f"• Centro de Eventos: {'Sí' if user.acceso_cde else 'No'}\n\n"
                        f"Este usuario requiere su aprobación para acceder al sistema.\n\n"
                        f"Saludos cordiales,\n"
                        f"Equipo de Gestor CCD"
                    )
                    try:
                        send_mail(
                            subject,
                            message,
                            settings.DEFAULT_FROM_EMAIL,
                            list(admin_emails),
                            fail_silently=False,
                        )
                    except Exception as e:
                        messages.warning(request, "El usuario se creó correctamente, pero no se pudo enviar la notificación a los administradores.")

                messages.success(request, f"Usuario '{user.username}' registrado exitosamente.")
                return redirect('libreria:inicio')

            except IntegrityError:
                messages.error(request, "Error: El nombre de usuario o correo electrónico ya existe en el sistema.")
            except ValidationError as e:
                messages.error(request, f"Error de validación: {', '.join(e.messages)}")
            except Exception as e:
                messages.error(request, f"Error inesperado al crear el usuario: {str(e)}")
        else:
            # Mensajes de error específicos por campo
            for field, errors in form.errors.items():
                field_label = form.fields[field].label if field in form.fields else field

                if 'username' in field:
                    for error in errors:
                        if 'unique' in error:
                            messages.error(request, "Este nombre de usuario ya está en uso. Por favor elija otro.")
                        else:
                            messages.error(request, f"Nombre de usuario: {error}")

                elif 'email' in field:
                    for error in errors:
                        if 'unique' in error:
                            messages.error(request, "Este correo electrónico ya está registrado.")
                        elif 'invalid' in error:
                            messages.error(request, "Ingrese una dirección de correo electrónico válida.")
                        else:
                            messages.error(request, f"Correo electrónico: {error}")

                elif 'password' in field:
                    for error in errors:
                        if 'too short' in error.lower():
                            messages.error(request, "La contraseña debe tener al menos 8 caracteres.")
                        elif 'too common' in error.lower():
                            messages.error(request, "La contraseña es demasiado común o insegura.")
                        else:
                            messages.error(request, f"Contraseña: {error}")

                else:
                    for error in errors:
                        messages.error(request, f"{field_label}: {error}")
    else:
        form = CustomUserForm()

    return render(request, 'usuario/crear_usuario.html', {
        'form': form,
        'admin_exists': admin_exists,
        'breadcrumbs': breadcrumbs
    })



def ver_usuario(request, user_id):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Ver usuario', 'url': reverse('libreria:ver_usuario', kwargs={'user_id': user_id})},
    ]
    usuario = get_object_or_404(CustomUser, id=user_id)
    return render(request, 'usuario/ver_perfil.html', {'usuario': usuario, 'breadcrumbs': breadcrumbs})

from django.urls import NoReverseMatch

def editar_usuario(request, user_id):
    usuario = get_object_or_404(CustomUser, id=user_id)

    # Intentar generar las URLs del breadcrumb de forma segura
    try:
        ver_usuario_url = reverse('libreria:ver_usuario', kwargs={'user_id': user_id})
    except NoReverseMatch:
        ver_usuario_url = '#'
        messages.warning(request, "No se pudo generar el enlace a 'Ver usuario'.")

    try:
        editar_usuario_url = reverse('libreria:editar_usuario', kwargs={'user_id': user_id})
    except NoReverseMatch:
        editar_usuario_url = '#'
        messages.warning(request, "No se pudo generar el enlace a 'Editar usuario'.")

    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Ver usuario', 'url': ver_usuario_url},
        {'name': 'Editar usuario', 'url': editar_usuario_url},
    ]

    if request.method == 'POST':
        form = CustomUserEditForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(request, "Usuario actualizado correctamente.")
            return redirect('libreria:inicio')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.capitalize()}: {error}")
    else:
        form = CustomUserEditForm(instance=usuario)

    return render(request, 'usuario/editar_usuario.html', {
        'form': form,
        'usuario': usuario,
        'breadcrumbs': breadcrumbs
    })
def verificar_usuario(request):
    usuarios = CustomUser.objects.all().order_by('id')
    context = {
        'usuarios': usuarios
    }
def lista_usuarios(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Listado de usuarios', 'url': '/lista_usuarios'},
    ]
    q = request.GET.get('q', '')
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    fecha_inicio = None
    fecha_fin = None

    usuarios = CustomUser.objects.all()
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        html = render_to_string('partials/tabla_usuarios.html', {'usuarios': usuarios})
        return JsonResponse({'html': html})
    # Filtrado por texto
    if q:
        usuarios = usuarios.filter(
            Q(username__icontains=q) |
            Q(email__icontains=q) |
            Q(role__icontains=q) |
            Q(area__icontains=q) |
            Q(cargo__icontains=q)
        )

    # Filtrado por fecha
    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            usuarios = usuarios.filter(fecha_registro__gte=fecha_inicio)
        except ValueError:
            usuarios = CustomUser.objects.none()

    if fecha_fin_str:
        try:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
            usuarios = usuarios.filter(fecha_registro__lte=fecha_fin)
        except ValueError:
            usuarios = CustomUser.objects.none()

    # Procesar actualización de permisos si es POST
    if request.method == 'POST' and 'actualizar_permisos' in request.POST:
        usuario_id = request.POST.get('usuario_id')
        try:
            usuario = CustomUser.objects.get(id=usuario_id)
            if usuario.role != 'Administrador':  # Solo actualizar si no es admin
                usuario.acceso_pap = 'acceso_pap' in request.POST
                usuario.acceso_caf = 'acceso_caf' in request.POST
                usuario.acceso_cde = 'acceso_cde' in request.POST
                usuario.save()
                messages.success(request, f"Permisos de {usuario.username} actualizados correctamente.")
            else:
                messages.info(request, "Los administradores tienen todos los permisos por defecto.")
        except CustomUser.DoesNotExist:
            messages.error(request, "Usuario no encontrado.")
        return redirect('libreria:lista_usuarios')

    # Paginación
    paginator = Paginator(usuarios, 4)
    page = request.GET.get('page')
    usuarios = paginator.get_page(page)

    return render(request, 'usuario/lista_usuarios.html', {
        'usuarios': usuarios,
        'breadcrumbs': breadcrumbs
    })
from django.contrib import messages
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse
from .models import CustomUser
import json
from django.views.decorators.http import require_http_methods

@login_required
def cambiar_estado_usuario(request, user_id):
    if request.method == 'POST':
        usuario = get_object_or_404(CustomUser, id=user_id)
        estado_anterior = usuario.is_active
        nuevo_estado = 'is_active' in request.POST
        
        # Evita que el usuario se desactive a sí mismo
        if request.user.id == usuario.id and not nuevo_estado:
            messages.error(request, "Error: No puedes desactivar tu propio usuario mientras estás autenticado.")
        # Evita que los administradores se desactiven a sí mismos
        elif request.user.id == usuario.id and usuario.role == 'Administrador' and not nuevo_estado:
            messages.error(request, "Error: Los usuarios con rol de Administrador no pueden desactivarse a sí mismos.")
        elif estado_anterior != nuevo_estado:
            usuario.is_active = nuevo_estado
            usuario.save()
            
            if nuevo_estado:
                messages.success(request, f"Éxito: El usuario '{usuario.username}' ha sido activado correctamente.")
            else:
                messages.error(request, f"Advertencia: El usuario '{usuario.username}' ha sido desactivado.")
        else:
            messages.error(request, f"Información: El estado del usuario '{usuario.username}' no cambió.")
    
    return redirect(reverse("libreria:lista_usuarios"))
from django.contrib.auth.hashers import check_password
import logging
logger = logging.getLogger(__name__)
@require_http_methods(["POST"])
def verificar_contraseña_actual(request):
    """
    Vista para verificar si la contraseña actual ingresada es correcta
    """
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Usuario no autenticado'}, status=401)
    
    try:
        data = json.loads(request.body)
        contraseña_actual = data.get('contraseña_actual', '')
        
        # Debug: Log para ver qué se está recibiendo
        logger.debug(f"Usuario: {request.user.username}")
        logger.debug(f"Contraseña recibida: {contraseña_actual}")
        
        # Método 1: Usar check_password directamente (más directo)
        password_valid = check_password(contraseña_actual, request.user.password)
        
        # Método 2: Como respaldo, usar authenticate
        user_auth = authenticate(username=request.user.username, password=contraseña_actual)
        
        # Debug: Log de resultados
        logger.debug(f"check_password result: {password_valid}")
        logger.debug(f"authenticate result: {user_auth is not None}")
        
        # Usar check_password como método principal
        if password_valid:
            return JsonResponse({'valida': True})
        else:
            # Intentar con authenticate como respaldo
            if user_auth is not None:
                return JsonResponse({'valida': True})
            else:
                return JsonResponse({
                    'valida': False, 
                    'mensaje': 'Contraseña actual incorrecta',
                    'debug': f'Usuario: {request.user.username}'
                })
            
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Datos JSON inválidos'}, status=400)
    except Exception as e:
        logger.error(f"Error en verificar_contraseña_actual: {str(e)}")
        return JsonResponse({'error': f'Error interno del servidor: {str(e)}'}, status=500)

def cambiar_contraseña(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Cambiar Contraseña', 'url': reverse('libreria:cambiar_contraseña')},
    ]
    if request.method == 'POST':
        form = CustomPasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            return redirect('libreria:inicio')  # Redirige después de cambiar la contraseña
    else:
        form = CustomPasswordChangeForm(user=request.user)
    return render(request, 'usuario/cambiar_contraseña.html', {'form': form, 'breadcrumbs': breadcrumbs})

def cambiar_contraseña_id(request, user_id):
    """Vista para cambiar contraseña de cualquier usuario por ID"""
    
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Cambiar Contraseña por ID', 'url': reverse('libreria:cambiar_contraseña_id', args=[user_id])},
    ]

    target_user = get_object_or_404(User, id=user_id)

    if request.method == 'POST':
        form = CustomPasswordChangeForm(user=target_user, data=request.POST)
        if form.is_valid():
            new_password = form.cleaned_data['new_password1']
            target_user.set_password(new_password)
            target_user.save()
            messages.success(request, f'Contraseña cambiada exitosamente para {target_user.username} (ID: {target_user.id})')
            return redirect('libreria:lista_usuarios', user_id=user_id)
        else:
            messages.error(request, 'Formulario inválido. Revisa los datos ingresados.')
    else:
        form = CustomPasswordChangeForm(user=target_user)

    all_users = User.objects.all().values(
    'id', 'username', 'email', 'role', 'area', 'cargo', 'is_active'
)

    context = {
        'form': form,
        'breadcrumbs': breadcrumbs,
        'all_users': all_users,
        'user_id': user_id,
        'user': target_user,
    }

    return render(request, 'usuario/cambiar_contraseña_id.html', context)
@require_http_methods(["GET"])
def verificar_contraseña(request):
    """
    Vista para verificar si la contraseña ya existe en la base de datos
    """
    password = request.GET.get('password', '')
    
    if not password:
        return JsonResponse({'error': 'No se proporcionó contraseña'}, status=400)
    
    try:
        # Verificar contra todos los usuarios
        usuarios = User.objects.all()
        password_existe = False
        
        for usuario in usuarios:
            if check_password(password, usuario.password):
                password_existe = True
                break
        
        return JsonResponse({
            'password_existe': password_existe,
            'mensaje': 'Esta contraseña ya está en uso. Por seguridad, elige una diferente.' if password_existe else 'Contraseña válida'
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Error al verificar contraseña: {str(e)}'}, status=500)
# views.py
from django.shortcuts import render, redirect
from django.contrib.auth import get_user_model
from django.contrib.auth.tokens import default_token_generator
from django.contrib.auth.forms import SetPasswordForm
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.template.loader import render_to_string
from django.core.mail import EmailMultiAlternatives
from django.urls import reverse
from django.conf import settings
import socket
# Obtener el modelo de usuario personalizado
User = get_user_model()

def password_reset_request(request):
    if request.method == "POST":
        email = request.POST.get("email")

        try:
            user = User.objects.get(email=email)
            token = default_token_generator.make_token(user)
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            reset_link = request.build_absolute_uri(
                reverse("libreria:password_reset_confirm", kwargs={"uidb64": uid, "token": token})
            )

            # Contexto del correo
            context = {
                'user': user,
                'reset_link': reset_link,
                'site_name': 'Gestor CCD',
                'company_name': 'Gestor CCD',
            }

            subject = "Restablecer tu contraseña - Gestor CCD"
            html_message = render_to_string("password_reset_email.html", context)

            text_message = f"""
Hola {user.username},

Recibimos una solicitud para restablecer la contraseña de tu cuenta en Gestor CCD.

Para restablecer tu contraseña, copia y pega el siguiente enlace en tu navegador:
{reset_link}

Si no solicitaste este cambio, puedes ignorar este mensaje.

El enlace será válido por 24 horas.

Saludos,
El equipo de Gestor CCD
"""

            try:
                # Enviar email
                msg = EmailMultiAlternatives(
                    subject,
                    text_message,
                    settings.DEFAULT_FROM_EMAIL,
                    [user.email]
                )
                msg.attach_alternative(html_message, "text/html")
                msg.send()

                return redirect(reverse("libreria:password_reset_done") + "?sent=true")

            except socket.timeout:
                # Error por tiempo de espera
                messages.error(request, "Hubo un problema de conexión al enviar el correo. Intenta nuevamente más tarde.")
                return render(request, "password_reset.html")

            except Exception as e:
                # Otro error (no se muestra al usuario por seguridad)
                print(f"Error enviando email: {e}")
                return redirect(reverse("libreria:password_reset_done") + "?sent=error")

        except User.DoesNotExist:
            # Usuario no encontrado, redirige sin indicar si existe o no
            return redirect(reverse("libreria:password_reset_done") + "?sent=notfound")

    return render(request, "password_reset.html")

def password_reset_done(request):
    sent_status = request.GET.get('sent', 'unknown')
    context = {
        'sent_status': sent_status,
        'show_success': sent_status == 'true',
        'show_error': sent_status == 'error',
        'show_not_found': sent_status == 'notfound'
    }
    return render(request, "password_reset_done.html", context)

def password_reset_confirm(request, uidb64, token):
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None
    
    if user is not None and default_token_generator.check_token(user, token):
        if request.method == "POST":
            form = SetPasswordForm(user, request.POST)
            if form.is_valid():
                form.save()
                return redirect(reverse("libreria:password_reset_complete"))
        else:
            form = SetPasswordForm(user)
        return render(request, "password_reset_confirm.html", {
            "form": form, 
            "user": user,
            "valid_link": True
        })
    else:
        return render(request, "password_reset_confirm.html", {
            "error": "El enlace no es válido o ha expirado.",
            "valid_link": False
        })

def password_reset_complete(request):
    return render(request, "password_reset_complete.html")
#VALIDAR INFORMACION

def validar_datos(request):
    email = request.GET.get('email', None)
    
    errores = {}

    # Validar correo electrónico
    if email and CustomUser.objects.filter(email=email).exists():
        errores['email'] = 'El email ya está en uso.'

    # Retornar los errores (si los hay) o una respuesta de validación exitosa
    return JsonResponse(errores if errores else {'valid': True})

def validate_password(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        password = data.get('password')
        user = request.user
        valid = user.check_password(password)
        return JsonResponse({'valid': valid})
    return JsonResponse({'valid': False})

from io import BytesIO
from datetime import datetime
from django.http import HttpResponse
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from django.contrib.staticfiles import finders
from papeleria.models import Articulo
def draw_table_on_canvas(canvas, doc):
    # Marca de agua
    watermark_path = finders.find('imagen/LOGO.png')
    if watermark_path:
        canvas.saveState()
        canvas.setFillColor(colors.Color(1, 1, 1, alpha=0.3))
        canvas.setStrokeColor(colors.Color(1, 1, 1, alpha=0.3))
        canvas.drawImage(watermark_path,
                         x=(doc.pagesize[0] - 600) / 2,
                         y=(doc.pagesize[1] - 600) / 2,
                         width=600, height=600, mask='auto')
        canvas.restoreState()


#GRAFICAS
def graficas_usuarios_activos(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Estadísticas', 'url': reverse('libreria:index_estadistica')}, 
        {'name': 'Gráfico de usuarios activos', 'url': reverse('libreria:graficas_usuarios_activos')}, 
    ]
    
    activos = CustomUser.objects.filter(is_active=True).count()
    inactivos = CustomUser.objects.filter(is_active=False).count()

    nombres = ['Activos', 'Inactivos']
    cantidades = [activos, inactivos]

    return render(request, 'estadisticas/grafica_usuarios.html', {
        'nombres': nombres,
        'cantidades': cantidades,
        'breadcrumbs': breadcrumbs
    })



def sesion_expirada(request):
    return render(request, 'sesion_expirada.html')

#PDF Y XSLS DE BAJO STOCK
def obtener_usuarios(request):
    query = request.GET.get('q')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    usuarios = CustomUser.objects.all()  # mejor nombrar 'usuarios' para que coincida

    if query:
        usuarios = usuarios.filter(
            Q(username__icontains=query) |
            Q(email__icontains=query) |
            Q(role__icontains=query) |
            Q(area__icontains=query) |
            Q(cargo__icontains=query)
        )

    if fecha_inicio and fecha_fin:
        usuarios = usuarios.filter(fecha_registro__range=[fecha_inicio, fecha_fin])

    return usuarios

def wrap_text(text, max_len=20):
    parts = [text[i:i+max_len] for i in range(0, len(text), max_len)]
    for i in range(len(parts) - 1):
        parts[i] += '-'  # Agrega guion al final de todas menos la última
    return '\n'.join(parts)

from reportlab.lib.enums import TA_CENTER
from reportlab.lib.styles import ParagraphStyle
def reporte_usuario_pdf(request):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                            leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)

    doc.title = "Listado de usuarios CCD"
    doc.author = "Gestor CCD"
    doc.subject = "Listado de usuarios CCD"
    doc.creator = "Sistema de Gestión CCD"

    elements = []
    styles = getSampleStyleSheet()

    # Título
    titulo = Paragraph("REPORTE DE USUARIOS CCD", styles["Title"])
    elements.append(titulo)

    # Encabezado empresa
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    encabezado_data = [
        ["GESTOR CCD", "Lista de artículos", "Correo:", f"Fecha: {fecha_actual}"],
        ["Cámara de comercio de Duitama", "Nit: 123456789", "contacto@gestorccd.com", "Teléfono: (123) 456-7890"],
    ]
    tabla_encabezado = Table(encabezado_data, colWidths=[180, 180, 180, 180])
    estilo_encabezado = TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    tabla_encabezado.setStyle(estilo_encabezado)
    elements.append(tabla_encabezado)

    # Tabla usuario autenticado
    usuario = request.user
    data_usuario = [["Usuario", "Email", "Rol", "Cargo"]]

    if usuario.is_authenticated:
        data_usuario.append([
            usuario.username,
            usuario.email,
            getattr(usuario, 'role', 'No definido'),
            getattr(usuario, 'cargo', 'No definido'),
        ])
    else:
        data_usuario.append(["Invitado", "-", "-", "-"])

    table_usuario = Table(data_usuario, colWidths=[180, 180, 180, 180])
    style_usuario = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table_usuario.setStyle(style_usuario)
    elements.append(table_usuario)

    # Obtener usuarios filtrados
    usuarios_filtrados = obtener_usuarios(request)

    if not usuarios_filtrados.exists():
        centered_style = ParagraphStyle(
        name="CenteredNormal",
        parent=styles["Normal"],
        alignment=TA_CENTER,)
        no_results = Paragraph("No se encontraron usuarios.", centered_style)
        elements.append(no_results)
    else:
        data_usuarios = [["ID", "Usuario", "Rol", "Correo", "Cargo", "Área", "Estado"]]
        for u in usuarios_filtrados:
            data_usuarios.append([
                wrap_text(str(u.id)),
                wrap_text(u.username),
                wrap_text(u.role),
                wrap_text(u.email),
                wrap_text(u.cargo),
                wrap_text(getattr(u, 'area', 'No definido')),
                wrap_text("Activo" if u.is_active else "Inactivo")
            ])

        tabla_articulos = Table(data_usuarios, colWidths=[30, 100, 100, 160, 160, 100, 70])
        style_articulos = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        tabla_articulos.setStyle(style_articulos)
        elements.append(tabla_articulos)

    # Construir el PDF
    doc.build(elements, onFirstPage=draw_table_on_canvas, onLaterPages=draw_table_on_canvas)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Listado de usuarios CCD.pdf"'
    return response

def reporte_usuario_excel(request):
    # Obtener usuarios filtrados
    usuarios = obtener_usuarios(request)

    # Crear archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Listado de Artículos CCD"

    # Configurar columnas
    column_widths = [10, 30, 20, 30, 30, 20, 15]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    ws.row_dimensions[1].height = 60
    ws.row_dimensions[2].height = 30

    # Título
    ws.merge_cells('A1:G1')
    ws['A1'] = "GESTOR CCD"
    ws['A1'].font = Font(size=24, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Subtítulo
    ws.merge_cells('A2:G2')
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    ws['A2'] = f"Listado de Artículos - {fecha_actual}"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    # Encabezados
    headers = ["ID", "Usuario", "Rol", "Correo", "Cargo", "Área", "Estado"]
    ws.append(headers)

    header_fill = PatternFill(start_color="FF0056B3", end_color="FF0056B3", fill_type="solid")
    for col in range(1, 8):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Estilo de bordes
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Agregar datos o mensaje de "no encontrados"
    if not usuarios.exists():
        ws.merge_cells('A4:G4')
        cell = ws['A4']
        cell.value = "No se encontraron usuarios."
        cell.font = Font(bold=False)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    else:
        for usuario in usuarios:
            estado = "Activo" if usuario.is_active else "Inactivo"
            ws.append([
                usuario.id,
                wrap_text(usuario.username),
                wrap_text(getattr(usuario, 'role', '')),
                wrap_text(usuario.email),
                wrap_text(getattr(usuario, 'cargo', '')),
                wrap_text(getattr(usuario, 'area', '')),
                estado
            ])

    # Estilo para títulos y subtítulo
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=7):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Estilo para encabezados y datos
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Filtro automático
    ws.auto_filter.ref = "A3:G3"

    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Listado de usuarios CCD.xlsx"'
    wb.save(response)
    return response