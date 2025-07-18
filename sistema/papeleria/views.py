from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib.sessions.models import Session
from django.contrib import messages
from django.urls import reverse
import os
from django.utils.timezone import localtime, make_aware
from django.core.paginator import Paginator
from django.utils.timezone import localtime
from datetime import datetime, timedelta
from django.utils.dateparse import parse_date
from django.db.models import Q, Sum, Count
from django.http import JsonResponse, HttpResponse
from django.template.loader import get_template
from django.db import transaction
from django.utils.timezone import now
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from django.core.mail import send_mail
from papeleria.models import Articulo, PedidoArticulo, Pedido
from papeleria.forms import LoginForm, ArticuloForm, ArticuloEditForm, PedidoArticuloForm
from libreria.models import CustomUser
from io import BytesIO
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from datetime import datetime
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from django.contrib.staticfiles import finders
from xhtml2pdf import pisa
User = get_user_model()
#ACCESO DENEGADO
#INDEX DE PAPELERIA
def error_404_view(request, exception):
    return render(request, 'acceso_denegado.html', status=404)
def timeouterror(request):
    try:
        # Simulación de una operación que puede causar un TimeoutError
        # Aquí va tu lógica real, como una conexión a red, base de datos externa, etc.
        raise TimeoutError("Error de tiempo de espera")  # Simulación

        # Si no ocurre error, puedes devolver otro template si lo deseas
        return render(request, 'exito.html')

    except TimeoutError:
        # Solo captura TimeoutError y redirige a lan_error.html
        return render(request, 'lan_error.html')
    
@never_cache
@login_required(login_url='/acceso_denegado/')
def index_pap(request):
    breadcrumbs = [{'name': 'Inicio', 'url': '/index_pap'}]

    # Aquí defines las condiciones para mostrar la alerta
    mostrar_alerta = True  # o alguna lógica
    bajo_stock = True      # o algún cálculo real
    es_papeleria = True    # esta variable ayuda a limitarlo al módulo

    context = {
        'breadcrumbs': breadcrumbs,
        'es_papeleria': es_papeleria,
    }

    return render(request, 'index_pap/index_pap.html', context)


#LOGIN  Y LOGOUT DE PAPELERIA
# Ejemplo para el login de papelería

def login_papeleria(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        user = authenticate(request, email=email, password=password)

        if user is not None:
            if user.is_active and getattr(user, 'acceso_pap', False):
                login(request, user)
                messages.success(request, "¡Inicio de sesión exitoso!")
                return redirect('papeleria:index_pap')
            else:
                messages.error(request, "No tienes permiso para acceder a este módulo.")
        else:
            messages.error(request, "Credenciales inválidas.")

    return render(request, 'login_pap/login_pap.html')

@never_cache
@login_required(login_url='/acceso_denegado/')
def logout_view(request):
    request.session.flush()  # borra toda la sesión y regenera session_key
    messages.success(request, "Has cerrado sesión correctamente.")
    response = redirect('libreria:inicio')
    response.delete_cookie('sessionid')
    return response
#VIEWS DE ARTICULOS
@login_required(login_url='/acceso_denegado/')
def crear_articulo(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Crear artículo', 'url': reverse('papeleria:crear_articulo')},
    ]

    if request.method == 'POST':
        form = ArticuloForm(request.POST)
        if form.is_valid():
            try:
                articulo = form.save(commit=False)
                articulo.registrado_por = request.user
                articulo.save()
                messages.success(request, '¡Artículo creado exitosamente!')
                return redirect('papeleria:listar_articulo')
            except Exception as e:
                messages.error(request, f'Error al crear el artículo: {str(e)}')
        else:
            # Mostrar errores específicos del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'Error en {field}: {error}')
    else:
        form = ArticuloForm()

    return render(request, 'articulo/crear_articulo.html', {
        'form': form,
        'breadcrumbs': breadcrumbs
    })
@login_required(login_url='/acceso_denegado/')
def editar_articulo(request, articulo_id):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Lista de articulos', 'url': reverse('papeleria:listar_articulo')}, 
        {'name': 'Editar artículo', 'url': reverse('papeleria:editar_articulo', kwargs={'articulo_id': articulo_id})},
    ]

    articulo = get_object_or_404(Articulo, id=articulo_id)

    if request.method == 'POST':
        form = ArticuloEditForm(request.POST, instance=articulo)
        if form.is_valid():
            try:
                # Guardar el artículo y registrar quién lo editó
                articulo_editado = form.save(commit=False)
                articulo_editado.registrado_por = request.user
                articulo_editado.save()
                
                messages.success(request, f'¡El artículo "{articulo.nombre}" actualizado correctamente!')
                return redirect('papeleria:listar_articulo')
            except Exception as e:
                messages.error(request, f'Error al guardar los cambios: {str(e)}')
        else:
            # Mostrar errores específicos del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'Error en {field}: {error}')
    else:
        form = ArticuloEditForm(instance=articulo)

    return render(request, 'articulo/editar_articulo.html', {
        'form': form,
        'articulo': articulo,
        'breadcrumbs': breadcrumbs
    })
@login_required(login_url='/acceso_denegado/')
def listar_articulo(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Listar artículos', 'url': reverse('papeleria:listar_articulo')},
    ]

    query = request.GET.get('q', '').strip()
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    articulos = Articulo.objects.select_related('registrado_por').order_by('nombre')

    # Búsqueda por texto
    if query:
        articulos = articulos.filter(
            Q(id__icontains=query) |
            Q(nombre__icontains=query) |
            Q(marca__icontains=query) |
            Q(observacion__icontains=query) |
            Q(tipo__icontains=query) |
            Q(precio__icontains=query) |
            Q(proveedor__icontains=query) |
            Q(cantidad__icontains=query) |
            Q(registrado_por__username__icontains=query)
        )

    # Filtrado por rango de fechas
    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            articulos = articulos.filter(fecha_registro__gte=fecha_inicio)
        except ValueError:
            articulos = Articulo.objects.none()

    if fecha_fin_str:
        try:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date() + timedelta(days=1)
            articulos = articulos.filter(fecha_registro__lt=fecha_fin)
        except ValueError:
            articulos = Articulo.objects.none()

    # Paginación
    paginator = Paginator(articulos, 4)
    page_number = request.GET.get('page')
    articulos_page = paginator.get_page(page_number)

    # Crear string con parámetros de búsqueda para la paginación
    query_params = ''
    if query:
        query_params += f'&q={query}'
    if fecha_inicio_str:
        query_params += f'&fecha_inicio={fecha_inicio_str}'
    if fecha_fin_str:
        query_params += f'&fecha_fin={fecha_fin_str}'  # Mostrar la fecha original (sin timedelta)

    return render(request, 'articulo/listar_articulo.html', {
        'articulos': articulos_page,
        'breadcrumbs': breadcrumbs,
        'query_params': query_params,
        'current_query': query,
        'current_fecha_inicio': fecha_inicio_str,
        'current_fecha_fin': fecha_fin_str,
    })

def buscar_articulo(request):
    query = request.GET.get('q', '').strip()
    if query:
        articulos = Articulo.objects.filter(nombre__icontains=query).values(
            'id', 'nombre', 'marca', 'observacion', 'precio', 'registrado_por', 'fecha_formateada', 'proveedor'
        )
        return JsonResponse(list(articulos), safe=False)
    return JsonResponse([], safe=False)

def eliminar_articulo(request, id):
    articulo = Articulo.objects.get(id=id)
    if request.method == 'POST':
        articulo.delete()
        messages.success(request, "Artículo eliminado correctamente.")
    return redirect('papeleria:listar_articulo')
@login_required(login_url='/acceso_denegado/')
def lista_stock_bajo(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Stock bajo', 'url': reverse('papeleria:lista_bajo_stock')}, 
    ]

    query = request.GET.get('q', '').strip()
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    # Filtrar solo artículos con cantidad menor a 10
    articulos_bajo_stock = Articulo.objects.filter(cantidad__lt=10)

    # Filtro por búsqueda de texto
    if query:
        articulos_bajo_stock = articulos_bajo_stock.filter(
            Q(nombre__icontains=query) |
            Q(marca__icontains=query) |
            Q(tipo__icontains=query) |
            Q(id__icontains=query) |
            Q(proveedor__icontains=query)
        )

    # Filtro por rango de fechas en fecha_registro
    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            articulos_bajo_stock = articulos_bajo_stock.filter(fecha_registro__gte=fecha_inicio)
        except ValueError:
            articulos_bajo_stock = Articulo.objects.none()

    if fecha_fin_str:
        try:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date() + timedelta(days=1)
            articulos_bajo_stock = articulos_bajo_stock.filter(fecha_registro__lt=fecha_fin)
        except ValueError:
            articulos_bajo_stock = Articulo.objects.none()

    bajo_stock = articulos_bajo_stock.exists()
    nombres_bajo_stock = [art.nombre for art in articulos_bajo_stock]

    # Paginación
    paginator = Paginator(articulos_bajo_stock, 4)
    page_number = request.GET.get('page')
    articulos_page = paginator.get_page(page_number)

    # Crear string con parámetros de búsqueda para la paginación
    query_params = ''
    if query:
        query_params += f'&q={query}'
    if fecha_inicio_str:
        query_params += f'&fecha_inicio={fecha_inicio_str}'
    if fecha_fin_str:
        query_params += f'&fecha_fin={fecha_fin_str}'  # Mostrar la fecha original (sin timedelta)

    return render(request, 'articulo/bajo_stock.html', {
        'articulos': articulos_page,
        'bajo_stock': bajo_stock,
        'nombres_bajo_stock': nombres_bajo_stock,
        'breadcrumbs': breadcrumbs,
        'query_params': query_params,
        'current_query': query,
        'current_fecha_inicio': fecha_inicio_str,
        'current_fecha_fin': fecha_fin_str,
    })
#VIEWS DE PEDIDOS

from django.shortcuts import render, redirect
from django.contrib import messages
from django.core.mail import send_mail
from django.urls import reverse
from django.conf import settings
from .models import Pedido, PedidoArticulo, Articulo
from django.contrib.auth import get_user_model
from django.utils import timezone
from django.contrib import messages
from django.conf import settings
from django.shortcuts import redirect, render
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.urls import reverse
from libreria.models import CustomUser
User = get_user_model()
@login_required(login_url='/acceso_denegado/')
def crear_pedido(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Realizar pedidos', 'url': reverse('papeleria:crear_pedido')}, 
    ]

    if request.method == 'POST':
        estado = 'Confirmado' if request.user.role == 'Administrador' else 'Pendiente'

        # ✅ OBTENER FECHA PERSONALIZADA SI SE ACTIVA EL CHECKBOX
        fecha_personalizada_str = request.POST.get('fecha_personalizada')
        if fecha_personalizada_str:
            try:
                fecha_personalizada = datetime.strptime(fecha_personalizada_str, '%Y-%m-%dT%H:%M')
                if timezone.is_naive(fecha_personalizada):
                    fecha_personalizada = make_aware(fecha_personalizada)
            except ValueError:
                fecha_personalizada = timezone.now()
        else:
            fecha_personalizada = timezone.now()

        pedido = Pedido.objects.create(
            registrado_por=request.user,
            estado=estado,
            fecha_estado=timezone.now() if estado == 'Confirmado' else None,
            fecha_pedido=fecha_personalizada  # ✅ APLICAMOS FECHA PERSONALIZADA O AUTOMÁTICA
        )

        articulos_ids = request.POST.getlist('articulo')
        tipos = request.POST.getlist('tipo_articulo')
        cantidades = request.POST.getlist('cantidad')

        area_usuario = request.user.area  # Usamos directamente el campo area del CustomUser

        for articulo_id, tipo, cantidad in zip(articulos_ids, tipos, cantidades):
            if articulo_id and tipo and cantidad:
                cantidad = int(cantidad)
                articulo = Articulo.objects.get(id=articulo_id)

                if estado == 'Confirmado' and articulo.cantidad < cantidad:
                    pedido.delete()
                    messages.error(request, f"No hay suficiente stock para el artículo: {articulo.nombre} (disponible: {articulo.cantidad}, solicitado: {cantidad})")
                    return redirect('papeleria:crear_pedido')

                if estado == 'Confirmado':
                    articulo.cantidad -= cantidad
                    articulo.save()

                PedidoArticulo.objects.create(
                    pedido=pedido,
                    articulo=articulo,
                    cantidad=cantidad,
                    tipo=tipo,
                    area=area_usuario,
                )

        admin_users = CustomUser.objects.filter(role='Administrador', is_active=True)
        admin_emails = [admin.email for admin in admin_users if admin.email]

        if admin_emails:
            admin_url = request.build_absolute_uri(reverse('papeleria:mis_pedidos'))
            context = {
                'usuario': request.user,
                'pedido': pedido,
                'admin_url': admin_url,
                'company_name': 'Gestor CCD',
                'articulos': PedidoArticulo.objects.filter(pedido=pedido),
                'fecha_pedido': localtime(pedido.fecha_pedido).strftime('%d/%m/%Y %H:%M')}

            html_message = render_to_string('pedidos/email_notificacion_pedido.html', context)
            
            # Usamos username en lugar de get_full_name()
            text_message = f"""
Hola Administrador,

Se ha registrado un nuevo pedido en el sistema de Papelería:

Usuario: {request.user.username}
Rol: {request.user.get_role_display()}  # Muestra el nombre legible del rol
Área: {request.user.get_area_display()}  # Muestra el nombre legible del área
Cargo: {request.user.cargo}
ID del Pedido: {pedido.id}
Estado: {estado}
Fecha: {pedido.fecha_pedido.strftime('%d/%m/%Y %H:%M')}

Detalle de artículos:
{'\n'.join([f"- {pa.articulo.nombre} x {pa.cantidad} ({pa.tipo})" for pa in PedidoArticulo.objects.filter(pedido=pedido)])}

Por favor revisa y gestiona este pedido en el sistema: {admin_url}

Este es un mensaje automático, por favor no respondas a este correo.
"""

            try:
                subject = f"Nuevo pedido {'confirmado' if estado == 'Confirmado' else 'pendiente'} - ID: {pedido.id}"
                msg = EmailMultiAlternatives(
                    subject,
                    text_message,
                    settings.DEFAULT_FROM_EMAIL,
                    admin_emails
                )
                msg.attach_alternative(html_message, "text/html")
                msg.send()

            except TimeoutError:
                messages.error(request, "El pedido fue creado, pero no se pudo enviar la notificación por correo debido a un problema de conexión (Timeout).")
            except Exception as e:
                print(f"Error enviando correo: {e}")

        messages.success(request, f"El pedido fue registrado correctamente con estado '{estado}'.")
        return redirect('papeleria:mis_pedidos' if request.user.role == 'Empleado' else 'papeleria:mis_pedidos')

    articulos = Articulo.objects.all()
    return render(request, 'pedidos/pedidos.html', {
        'articulos': articulos,
        'breadcrumbs': breadcrumbs,
    })
@login_required(login_url='/acceso_denegado/')
def mis_pedidos(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Listado de pedidos', 'url': reverse('papeleria:listado_pedidos')},
    ]
    
    query = request.GET.get('q', '').strip()
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    # Obtiene pedidos del usuario actual ordenados por fecha_pedido descendente
    pedidos = Pedido.objects.filter(registrado_por=request.user).order_by('-fecha_pedido')

    if query:
        pedidos = pedidos.filter(
            Q(registrado_por__username__icontains=query) |
            Q(estado__icontains=query) |
            Q(articulos__articulo__nombre__icontains=query) |
            Q(articulos__cantidad__icontains=query) |
            Q(articulos__tipo__icontains=query) |
            Q(articulos__area__icontains=query) |
            Q(id__icontains=query)
        ).distinct()

    # Manejo de fechas con timedelta
    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            pedidos = pedidos.filter(fecha_pedido__gte=fecha_inicio)
        except ValueError:
            pass
            
    if fecha_fin_str:
        try:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date() + timedelta(days=1)
            pedidos = pedidos.filter(fecha_pedido__lt=fecha_fin)
        except ValueError:
            pass

    paginator = Paginator(pedidos, 4)
    page_number = request.GET.get('page')
    pedidos_page = paginator.get_page(page_number)

    # Construir parámetros de búsqueda para paginación
    query_params = ''
    if query:
        query_params += f'&q={query}'
    if fecha_inicio_str:
        query_params += f'&fecha_inicio={fecha_inicio_str}'
    if fecha_fin_str:
        query_params += f'&fecha_fin={fecha_fin_str}'

    return render(request, 'pedidos/mis_pedidos.html', {
        'pedidos': pedidos_page,
        'breadcrumbs': breadcrumbs,
        'query_params': query_params,
        'current_query': query,
        'current_fecha_inicio': fecha_inicio_str,
        'current_fecha_fin': fecha_fin_str,
    })
from django.utils import timezone

@csrf_exempt
@require_POST
def cambiar_estado_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)

    if request.method == 'POST':
        nuevo_estado = request.POST.get('estado')
        
        if nuevo_estado:
            if nuevo_estado == 'confirmado':
                articulos_del_pedido = pedido.articulos.all()
                for articulo_pedido in articulos_del_pedido:
                    articulo = articulo_pedido.articulo
                    cantidad_pedida = articulo_pedido.cantidad

                    if articulo.cantidad >= cantidad_pedida:
                        articulo.cantidad -= cantidad_pedida
                        articulo.save()
                    else:
                        messages.error(request, f"No hay suficiente stock de {articulo.nombre}.")
                        return redirect('papeleria:pedidos_pendientes')

            pedido.estado = nuevo_estado
            pedido.fecha_actualizacion = datetime.now()
            pedido.save()

            usuario = pedido.registrado_por
            if usuario and usuario.email:
                articulos_lista = "\n".join(
                    f"<li class='articulo-item'>{articulo.cantidad} {articulo.articulo.nombre}</li>"
                    for articulo in pedido.articulos.all()
                )

                html_content = f"""
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Estado de Pedido Actualizado</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            line-height: 1.6;
            color: #333333;
            max-width: 600px;
            margin: 0 auto;
            padding: 20px;
            background-color: #f5f5f5;
        }}
        .email-container {{
            background-color: #ffffff;
            border-radius: 8px;
            box-shadow: 0 2px 10px rgba(0, 0, 0, 0.1);
            overflow: hidden;
        }}
        .header {{
            color: white;
            padding: 20px;
            text-align: center;
        }}
        .logo {{
            max-width: 380px;
            height: 200px;
            margin-bottom: 15px;
        }}
        .content {{
            padding: 25px;
        }}
        .status {{
            font-size: 18px;
            font-weight: bold;
            margin: 15px 0;
            padding: 10px;
            border-radius: 5px;
            text-align: center;
        }}
        .status.confirmado {{
            color: #155724;
        }}
        .status.cancelado {{
            background-color: #f8d7da;
            color: #721c24;
        }}
        .status.pendiente {{
            color: #856404;
        }}
        .info-box {{
            background-color: #f8f9fa;
            border-left: 4px solid #007bff;
            padding: 15px;
            margin: 20px 0;
        }}
        .info-label {{
            font-weight: bold;
            color: #495057;
            display: inline-block;
            min-width: 150px;
        }}
        .articulos-list {{
            margin: 15px 0;
            padding-left: 20px;
        }}
        .articulo-item {{
            margin-bottom: 8px;
        }}
        .footer {{
            text-align: center;
            padding: 15px;
            font-size: 12px;
            color: #6c757d;
            border-top: 1px solid #e9ecef;
            background-color: #f8f9fa;
        }}
    </style>
</head>
<body>
    <div class="email-container">
        <div class="header">
            <img src="https://ccduitama.org.co/wp-content/uploads/2021/05/LOGOCCD-TRANSPARENCIA.png" alt="Logo CCD" class="logo">
            <h2>Actualización de Estado de Pedido</h2>
        </div>
        
        <div class="content">
            <p>Hola <strong>{usuario.username}</strong>,</p>
            <p>Te informamos que el estado de tu pedido en el módulo de papeleria ha sido actualizado:</p>
            
            <div class="info-box">
                <p><span class="info-label">Número de Pedido:</span> #{pedido.id}</p>
                <p><span class="info-label">Fecha de Actualización:</span> {pedido.fecha_actualizacion.strftime('%d/%m/%Y %H:%M')}</p>
                <p><span class="info-label">Nuevo Estado:</span> <span class="status {pedido.estado.lower()}">{pedido.estado.upper()}</span></p>

                <h3>Detalle del Pedido:</h3>
            <ul class="articulos-list">
                {articulos_lista}
            </ul>
            </div>
            
            <p>Para más información, puedes acceder al sistema de gestión de pedidos.</p>
        </div>
        
        <div class="footer">
            <p>Este es un mensaje automático, por favor no respondas a este correo.</p>
            <p>© {datetime.now().year} Gestor CCD - Todos los derechos reservados</p>
        </div>
    </div>
</body>
</html>
                """

                text_content = f"""
Actualización de Estado de Pedido

Hola {usuario.username},

Te informamos que el estado de tu pedido ha sido actualizado:

Número de Pedido: #{pedido.id}
Fecha de Actualización: {pedido.fecha_actualizacion.strftime('%d/%m/%Y %H:%M')}
Nuevo Estado: {pedido.estado.upper()}

Detalle del Pedido:
{"".join([f"- {articulo.cantidad} x {articulo.articulo.nombre}\n" for articulo in pedido.articulos.all()])}

Para más información, puedes acceder al sistema de gestión de pedidos.

© {datetime.now().year} Gestor CCD - Todos los derechos reservados
                """

                try:
                    send_mail(
                        subject=f'[Pedido #{pedido.id}] Estado actualizado a {pedido.estado.upper()}',
                        message=text_content,
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        recipient_list=[usuario.email],
                        fail_silently=False,
                        html_message=html_content
                    )
                    messages.success(request, 'Estado del pedido actualizado y notificación enviada.')
                except TimeoutError:
                    messages.error(request, 'Estado actualizado, pero hubo un error al enviar el correo electrónico.')

            else:
                messages.success(request, 'Estado del pedido actualizado.')

            return redirect('papeleria:pedidos_pendientes')

    messages.error(request, 'No se pudo actualizar el estado del pedido.')
    return redirect('papeleria:pedidos_pendientes')

@login_required(login_url='/acceso_denegado/')
def listado_pedidos(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Listado de pedidos', 'url': reverse('papeleria:listado_pedidos')},
    ]

    query = request.GET.get('q', '').strip()
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    # Filtrar pedidos confirmados o cancelados
    pedidos = Pedido.objects.filter(estado__in=['Confirmado', 'Cancelado']).order_by('-fecha_estado')

    
    if query:
        pedidos = pedidos.filter(
            Q(registrado_por__username__icontains=query) |
            Q(estado__icontains=query) |
            Q(articulos__articulo__nombre__icontains=query) |
            Q(articulos__cantidad__icontains=query) |
            Q(articulos__tipo__icontains=query) |
            Q(articulos__area__icontains=query) |
            Q(id__icontains=query)
        ).distinct()

    # Manejo de fechas con timedelta
    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            pedidos = pedidos.filter(fecha_pedido__gte=fecha_inicio)
        except ValueError:
            pass
            
    if fecha_fin_str:
        try:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date() + timedelta(days=1)
            pedidos = pedidos.filter(fecha_pedido__lt=fecha_fin)
        except ValueError:
            pass

    paginator = Paginator(pedidos, 4)
    page_number = request.GET.get('page')
    pedidos_page = paginator.get_page(page_number)

    # Construir parámetros de búsqueda para paginación
    query_params = ''
    if query:
        query_params += f'&q={query}'
    if fecha_inicio_str:
        query_params += f'&fecha_inicio={fecha_inicio_str}'
    if fecha_fin_str:
        query_params += f'&fecha_fin={fecha_fin_str}'

    return render(request, 'pedidos/lista_pedidos.html', {
        'pedidos': pedidos_page,
        'breadcrumbs': breadcrumbs,
        'query_params': query_params,
        'current_query': query,
        'current_fecha_inicio': fecha_inicio_str,
        'current_fecha_fin': fecha_fin_str,
    })
@login_required(login_url='/acceso_denegado/')
def pedidos_pendientes(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Pedidos pendientes', 'url': reverse('papeleria:pedidos_pendientes')},
    ]

    query = request.GET.get('q', '').strip()
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')

    # Filtrar solo pedidos pendientes
    pedidos = Pedido.objects.filter(estado='Pendiente').order_by('-fecha_pedido')
    
    # Filtrado por búsqueda de texto
    if query:
        pedidos = pedidos.filter(
            Q(registrado_por__username__icontains=query) |
            Q(estado__icontains=query) |
            Q(articulos__articulo__nombre__icontains=query) |
            Q(articulos__cantidad__icontains=query) |
            Q(articulos__tipo__icontains=query) |
            Q(articulos__area__icontains=query) |
            Q(id__icontains=query)
        ).distinct()

    # Filtrado por fechas con manejo de errores
    try:
        if fecha_inicio_str:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            pedidos = pedidos.filter(fecha_pedido__gte=fecha_inicio)
            
        if fecha_fin_str:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date() + timedelta(days=1)
            pedidos = pedidos.filter(fecha_pedido__lt=fecha_fin)
    except ValueError:
        pass  # Mantener el queryset original si hay error en fechas

    # Paginación
    paginator = Paginator(pedidos, 4)
    page_number = request.GET.get('page')
    pedidos_page = paginator.get_page(page_number)

    # Construir parámetros de búsqueda para paginación
    query_params = ''
    if query:
        query_params += f'&q={query}'
    if fecha_inicio_str:
        query_params += f'&fecha_inicio={fecha_inicio_str}'
    if fecha_fin_str:
        query_params += f'&fecha_fin={fecha_fin_str}'

    return render(request, 'pedidos/confirmar_pedido.html', {
        'pedidos': pedidos_page,
        'breadcrumbs': breadcrumbs,
        'query_params': query_params,
        'current_query': query,
        'current_fecha_inicio': fecha_inicio_str,
        'current_fecha_fin': fecha_fin_str,
    })
#VALIDAR DATOS
def validar_datos(request):
    email = request.GET.get('email', None)
    
    errores = {}

    # Validar correo electrónico
    if email and CustomUser.objects. filter(email=email).exists():
        errores['email'] = 'El email ya está en uso.'

    # Retornar los errores (si los hay) o una respuesta de validación exitosa
    return JsonResponse(errores if errores else {'valid': True})

def verificar_nombre_articulo(request):
    nombre = request.GET.get('nombre', '')
    existe = Articulo.objects.filter(nombre=nombre).exists()
    return JsonResponse({'existe': existe})

#ESTADISTICAS 
@login_required(login_url='/acceso_denegado/')
def index_estadistica(request):
    breadcrumbs = [
    {'name': 'Inicio', 'url': '/index_pap'},
    {'name': 'Estadisticas', 'url': '/index_estadistica/'},  # URL fija
]

    return render(request, 'estadisticas/index_estadistica.html', {'breadcrumbs': breadcrumbs})
@login_required(login_url='/acceso_denegado/')
def estadisticas_articulos(request):
    articulos = Articulo.objects.all().order_by('-cantidad')  # Puedes ordenar por cantidad, nombre, etc.
    total_cantidad = articulos.aggregate(total=Sum('cantidad'))['total'] or 0
    return render(request, 'estadisticas/estadisticas_articulos.html', {
        'articulos': articulos,
        'total_cantidad': total_cantidad,
    })
#GRAFIC DE CANTIDAD DE ARTICULOS
@login_required(login_url='/acceso_denegado/')
def graficas_articulos(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Estadisticas', 'url': reverse('papeleria:index_estadistica')}, 
        {'name': 'Grafico de articulos', 'url': reverse('papeleria:graficas_articulos')}, 
    ]
    articulos = Articulo.objects.all()
    nombres = [art.nombre for art in articulos]
    cantidades = [art.cantidad for art in articulos]

    return render(request, 'estadisticas/grafica_articulos.html', {
        'nombres': nombres,
        'cantidades': cantidades,
        'breadcrumbs': breadcrumbs
    })
#grafica de usuario
@login_required(login_url='/acceso_denegado/')
def graficas_usuario(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Estadisticas', 'url': reverse('papeleria:index_estadistica')}, 
        {'name': 'Grafico de usuarios activos e inactivos', 'url': reverse('papeleria:graficas_usuarios')}, 
    ]
    activos = CustomUser.objects.filter(is_active=True).count()
    inactivos = CustomUser.objects.filter(is_active=False).count()

    nombres = ['Activos', 'Inactivos']
    cantidades = [activos, inactivos] 
    return render(request, 'estadisticas/grafica_usuarios.html', {'nombres':nombres,
                                                                  'cantidades': cantidades,
                                                                  'breadcrumbs': breadcrumbs})
@login_required(login_url='/acceso_denegado/')
def grafica_pedidos(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Estadísticas', 'url': reverse('papeleria:index_estadistica')},
        {'name': 'Gráfico de estado de pedidos', 'url': reverse('papeleria:grafica_pedidos')},
    ]

    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    pedidos = Pedido.objects.all()

    if fecha_inicio:
        pedidos = pedidos.filter(fecha_pedido__gte=fecha_inicio)
    if fecha_fin:
        pedidos = pedidos.filter(fecha_pedido__lte=fecha_fin)

    pendientes = pedidos.filter(estado='Pendiente').count()
    confirmados = pedidos.filter(estado='Confirmado').count()
    cancelados = pedidos.filter(estado='Cancelado').count()

    nombres = ['Pendiente', 'Confirmado', 'Cancelado']
    cantidades = [pendientes, confirmados, cancelados]

    return render(request, 'estadisticas/grafica_estado_pendiente.html', {
        'nombres': nombres,
        'cantidades': cantidades,
        'breadcrumbs': breadcrumbs,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin
    })

from django.shortcuts import render
from django.db.models import Sum
from django.urls import reverse
from .models import PedidoArticulo
@login_required(login_url='/acceso_denegado/')
def grafica_pedidos_administrativa(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Estadísticas', 'url': reverse('papeleria:index_estadistica')},
        {'name': 'Gráfico de pedidos Administrativa', 'url': reverse('papeleria:pedidos_administrativa')},
    ]

    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    pedidos = PedidoArticulo.objects.filter(area='Administrativa')

    if fecha_inicio:
        pedidos = pedidos.filter(pedido__fecha_pedido__gte=fecha_inicio)
    if fecha_fin:
        pedidos = pedidos.filter(pedido__fecha_pedido__lte=fecha_fin)

    pedidos_por_area_articulo = pedidos.values(
        'area',
        'articulo__nombre'
    ).annotate(
        total_cantidad=Sum('cantidad')
    ).order_by('articulo__nombre')

    etiquetas = [item['articulo__nombre'] or 'Sin artículo' for item in pedidos_por_area_articulo]
    cantidades = [item['total_cantidad'] or 0 for item in pedidos_por_area_articulo]

    return render(request, 'estadisticas/grafico_pedido_area.html', {
        'nombres': etiquetas,
        'cantidades': cantidades,
        'breadcrumbs': breadcrumbs,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin
    })
@login_required(login_url='/acceso_denegado/')
def grafica_pedidos_rues(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Estadísticas', 'url': reverse('papeleria:index_estadistica')},
        {'name': 'Pedidos RUES', 'url': reverse('papeleria:pedidos_rues')},
    ]

    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    pedidos = PedidoArticulo.objects.filter(area='Registros públicos')

    if fecha_inicio:
        pedidos = pedidos.filter(pedido__fecha_pedido__gte=fecha_inicio)
    if fecha_fin:
        pedidos = pedidos.filter(pedido__fecha_pedido__lte=fecha_fin)

    pedidos_por_area_articulo = pedidos.values(
        'area',
        'articulo__nombre'
    ).annotate(
        total_cantidad=Sum('cantidad')
    ).order_by('articulo__nombre')

    etiquetas = [item['articulo__nombre'] or 'Sin artículo' for item in pedidos_por_area_articulo]
    cantidades = [item['total_cantidad'] or 0 for item in pedidos_por_area_articulo]

    return render(request, 'estadisticas/grafico_pedido_rues.html', {
        'nombres': etiquetas,
        'cantidades': cantidades,
        'breadcrumbs': breadcrumbs,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin
    })
@login_required(login_url='/acceso_denegado/')
def grafica_pedidos_presidencia(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Estadísticas', 'url': reverse('papeleria:index_estadistica')},
        {'name': 'Pedidos Presidencia', 'url': reverse('papeleria:pedidos_presidencia')},
    ]

    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    pedidos = PedidoArticulo.objects.filter(area='Presidencia')

    if fecha_inicio:
        pedidos = pedidos.filter(pedido__fecha_pedido__gte=fecha_inicio)
    if fecha_fin:
        pedidos = pedidos.filter(pedido__fecha_pedido__lte=fecha_fin)

    pedidos_por_area_articulo = pedidos.values(
        'area',
        'articulo__nombre'
    ).annotate(
        total_cantidad=Sum('cantidad')
    ).order_by('articulo__nombre')

    etiquetas = [item['articulo__nombre'] or 'Sin artículo' for item in pedidos_por_area_articulo]
    cantidades = [item['total_cantidad'] or 0 for item in pedidos_por_area_articulo]

    return render(request, 'estadisticas/grafico_pedido_presidencia.html', {
        'nombres': etiquetas,
        'cantidades': cantidades,
        'breadcrumbs': breadcrumbs,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin
    })
@login_required(login_url='/acceso_denegado/')
@never_cache
def grafica_pedidos_financiera(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Estadísticas', 'url': reverse('papeleria:index_estadistica')},
        {'name': 'Pedidos financiera', 'url': reverse('papeleria:pedidos_financiera')},
    ]

    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    pedidos = PedidoArticulo.objects.filter(area='Financiera')

    if fecha_inicio:
        pedidos = pedidos.filter(pedido__fecha_pedido__gte=fecha_inicio)
    if fecha_fin:
        pedidos = pedidos.filter(pedido__fecha_pedido__lte=fecha_fin)

    pedidos_por_area_articulo = pedidos.values(
        'area',
        'articulo__nombre'
    ).annotate(
        total_cantidad=Sum('cantidad')
    ).order_by('articulo__nombre')

    etiquetas = [item['articulo__nombre'] or 'Sin artículo' for item in pedidos_por_area_articulo]
    cantidades = [item['total_cantidad'] or 0 for item in pedidos_por_area_articulo]

    return render(request, 'estadisticas/grafico_pedido_financiera.html', {
        'nombres': etiquetas,
        'cantidades': cantidades,
        'breadcrumbs': breadcrumbs,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin
    })
@login_required(login_url='/acceso_denegado/')
def grafica_pedidos_gestion_empresarial(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Estadísticas', 'url': reverse('papeleria:index_estadistica')},
        {'name': 'Pedidos gestión empresarial', 'url': reverse('papeleria:pedidos_gestion_empresarial')},

    ]

    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    pedidos = PedidoArticulo.objects.filter(area='Gestión empresarial')

    if fecha_inicio:
        pedidos = pedidos.filter(pedido__fecha_pedido__gte=fecha_inicio)
    if fecha_fin:
        pedidos = pedidos.filter(pedido__fecha_pedido__lte=fecha_fin)

    pedidos_por_area_articulo = pedidos.values(
        'area',
        'articulo__nombre'
    ).annotate(
        total_cantidad=Sum('cantidad')
    ).order_by('articulo__nombre')

    etiquetas = [item['articulo__nombre'] or 'Sin artículo' for item in pedidos_por_area_articulo]
    cantidades = [item['total_cantidad'] or 0 for item in pedidos_por_area_articulo]

    return render(request, 'estadisticas/grafico_pedido_gestion_empresarial.html', {
        'nombres': etiquetas,
        'cantidades': cantidades,
        'breadcrumbs': breadcrumbs,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin
    })
@login_required(login_url='/acceso_denegado/')
def grafica_pedidos_competitividad(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Estadísticas', 'url': reverse('papeleria:index_estadistica')},
        {'name': 'Pedidos gestión empresarial', 'url': reverse('papeleria:pedidos_gestion_empresarial')},

    ]

    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    pedidos = PedidoArticulo.objects.filter(area='Competitividad')

    if fecha_inicio:
        pedidos = pedidos.filter(pedido__fecha_pedido__gte=fecha_inicio)
    if fecha_fin:
        pedidos = pedidos.filter(pedido__fecha_pedido__lte=fecha_fin)

    pedidos_por_area_articulo = pedidos.values(
        'area',
        'articulo__nombre'
    ).annotate(
        total_cantidad=Sum('cantidad')
    ).order_by('articulo__nombre')

    etiquetas = [item['articulo__nombre'] or 'Sin artículo' for item in pedidos_por_area_articulo]
    cantidades = [item['total_cantidad'] or 0 for item in pedidos_por_area_articulo]

    return render(request, 'estadisticas/grafico_pedido_competitividad.html', {
        'nombres': etiquetas,
        'cantidades': cantidades,
        'breadcrumbs': breadcrumbs,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin
    })
@login_required(login_url='/acceso_denegado/')
def grafica_bajo_Stock(request):
    breadcrumbs = [
        {'name': 'Inicio', 'url': '/index_pap'},
        {'name': 'Estadisticas', 'url': reverse('papeleria:index_estadistica')}, 
        {'name': 'Grafico de bajo stock', 'url': reverse('papeleria:grafica_bajoStock')}, 
    ]
    articulos = Articulo.objects.filter(cantidad__lt=10)
    nombres = [art.nombre for art in articulos]
    cantidades = [art.cantidad for art in articulos]
    return render(request, 'estadisticas/grafica_bajoStock.html', {
        'nombres': nombres,
        'cantidades': cantidades,
        'breadcrumbs': breadcrumbs
    })

#PDF Y XSLS ARTICULOS
def wrap_text(text, max_len=20):
    parts = [text[i:i+max_len] for i in range(0, len(text), max_len)]
    for i in range(len(parts) - 1):
        parts[i] += '-'  # Agrega guion al final de todas menos la última
    return '\n'.join(parts)
#EXCLUSIVAMENTE PARA PEDIDOS 
def wrap_text_p(text, max_len=24
                ):
    parts = [text[i:i+max_len] for i in range(0, len(text), max_len)]
    for i in range(len(parts) - 1):
        parts[i] += '-'  # Agrega guion al final de todas menos la última
    return '\n'.join(parts)
def draw_table_on_canvas(canvas, doc):
    # Marca de agua
    watermark_path = finders.find('imagen/LOGO.png')
    if watermark_path:
        canvas.saveState()
        canvas.setFillColor(colors.Color(1, 1, 1, alpha=0.3))
        canvas.setStrokeColor(colors.Color(1, 1, 1, alpha=0.3))
        canvas.drawImage(watermark_path,
                         x=(doc.pagesize[0] - 600) / 2,
                         y=(doc.pagesize[1] - 600) / 2,
                         width=600, height=600, mask='auto')
        canvas.restoreState()
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.styles import ParagraphStyle
def get_articulos_filtrados(request):
    query = request.GET.get('q')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    articulos = Articulo.objects.select_related('registrado_por').order_by('nombre')


    if query:
        articulos = articulos.filter(
            Q(nombre__icontains=query) |
            Q(marca__icontains=query) |
            Q(observacion__icontains=query) |
            Q(tipo__icontains=query) |
            Q(precio__icontains=query) |
            Q(proveedor__icontains=query) |
            Q(cantidad__icontains=query) |
            Q(registrado_por__username__icontains=query)
        )

    if fecha_inicio and fecha_fin:
        articulos = articulos.filter(fecha_registro__range=[fecha_inicio, fecha_fin])

    return articulos


def reporte_articulo_pdf(request):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                            leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)

    doc.title = "Listado de Artículos CCD"
    doc.author = "CCD"
    doc.subject = "Listado de artículos"
    doc.creator = "Sistema de Gestión CCD"

    elements = []
    styles = getSampleStyleSheet()

    # Título
    titulo = Paragraph("REPORTE DE ARTÍCULOS", styles["Title"])
    elements.append(titulo)

    # Encabezado empresa
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    encabezado_data = [
        ["GESTOR CCD", "Lista de artículos", "Correo:", f"Fecha: {fecha_actual}"],
        ["Cámara de comercio de Duitama", "Nit: 123456789", "contacto@gestorccd.com", "Teléfono: (123) 456-7890"],
    ]
    tabla_encabezado = Table(encabezado_data, colWidths=[180, 180, 180, 180])
    estilo_encabezado = TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    tabla_encabezado.setStyle(estilo_encabezado)
    elements.append(tabla_encabezado)

    # Tabla usuario
    usuario = request.user
    data_usuario = [["Usuario", "Email", "Rol", "Cargo"]]
    data_usuario.append([
        usuario.username,
        usuario.email,
        getattr(usuario, 'role', 'No definido'),
        getattr(usuario, 'cargo', 'No definido'),
    ])
    table_usuario = Table(data_usuario, colWidths=[180, 180, 180, 180])
    style_usuario = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table_usuario.setStyle(style_usuario)
    elements.append(table_usuario)

    # Obtener artículos filtrados
    articulos_filtrados = get_articulos_filtrados(request)

    if not articulos_filtrados.exists():
        centered_style = ParagraphStyle(
            name="CenteredNormal",
            parent=styles["Normal"],
            alignment=TA_CENTER,
        )
        no_results = Paragraph("No se encontraron artículos.", centered_style)
        elements.append(no_results)
    else:
        # Tabla artículos
        data_articulos = [["ID", "Nombre", "Marca", "Tipo", "Precio", "Cantidad", "Observación"]]

        for articulo in articulos_filtrados:
            data_articulos.append([
                wrap_text(str(articulo.id)),
                wrap_text(articulo.nombre),
                wrap_text(articulo.marca),
                wrap_text(articulo.tipo),
                wrap_text("{:,}".format(articulo.precio)),
                wrap_text(str(articulo.cantidad)),
                wrap_text(articulo.observacion),
            ])

        tabla_articulos = Table(data_articulos, colWidths=[20, 140, 120, 110, 100, 80, 150])
        style_articulos = TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        tabla_articulos.setStyle(style_articulos)
        elements.append(tabla_articulos)

    # Marca de agua u otros elementos
    doc.build(elements, onFirstPage=draw_table_on_canvas, onLaterPages=draw_table_on_canvas)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Listado de articulos CCD.pdf"'
    return response
def reporte_articulo_excel(request):
    articulos = get_articulos_filtrados(request)

    # Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Listado de Artículos CCD"

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 33

    ws.row_dimensions[1].height = 60
    ws.row_dimensions[2].height = 30

    ws.merge_cells('A1:G1')
    ws['A1'] = "GESTOR CCD"
    ws['A1'].font = Font(size=24, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:G2')
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    ws['A2'] = f"Listado de Artículos {fecha_actual}"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    for row in [1, 2]:
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws[f"{col}{row}"].border = border
            ws[f"{col}{row}"].alignment = Alignment(horizontal='center', vertical='center')

    headers = ["ID", "Nombre", "Marca", "Tipo", "Precio", "Cantidad", "Observación"]
    ws.append(headers)
    header_fill = PatternFill(start_color="FF0056B3", end_color="FF0056B3", fill_type="solid")

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        cell = ws[f"{col}3"]
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for articulo in articulos:
        ws.append([
            wrap_text(str(articulo.id)),
            wrap_text(articulo.nombre),
            wrap_text(articulo.marca),
            wrap_text(articulo.tipo),
            wrap_text("{:,}".format(articulo.precio)),
            wrap_text(str(articulo.cantidad)),
            str((articulo.observacion)),
        ])

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Listado de articulos CCD.xlsx"'
    wb.save(response)
    return response


#PDF Y XSLS DE PEDIDOS
def get_pedidos_filtrados(request):
    query = request.GET.get('q')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    pedidos = Pedido.objects.filter(estado__in=['Confirmado', 'Cancelado']).order_by('-fecha_pedido')

    if query:
        pedidos = pedidos.filter(
            Q(registrado_por__username__icontains=query) |
            Q(estado__icontains=query) |
            Q(articulos__articulo__nombre__icontains=query) |
            Q(articulos__cantidad__icontains=query) |
            Q(articulos__tipo__icontains=query) |
            Q(articulos__area__icontains=query) |
            Q(id__icontains=query)
        ).distinct()

    if fecha_inicio and fecha_fin:
        try:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d") + timedelta(days=1)  # sumar un día para incluir todo el día fin
            pedidos = pedidos.filter(fecha_pedido__gte=fecha_inicio_dt, fecha_pedido__lt=fecha_fin_dt)
        except ValueError:
            pass  # puedes mostrar un mensaje de error si las fechas no son válidas

    return pedidos

def reporte_pedidos_pdf(request):
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=40,
        rightMargin=40,
        topMargin=40,
        bottomMargin=40
    )

    doc.title = "Listado de pedidos CCD"
    doc.author = "CCD"
    doc.subject = "Listado de pedidos"
    doc.creator = "Sistema de Gestión CCD"

    elements = []
    styles = getSampleStyleSheet()

    # Título principal
    titulo = Paragraph("REPORTE DE PEDIDOS CONFIRMADOS Y CANCELADOS", styles["Title"])
    elements.append(titulo)

    # Encabezado institucional
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    encabezado_data = [
        ["GESTOR CCD", "Lista de usuarios", "Correo: gestorccd@gmail.com", f"Fecha: {fecha_actual}"],
        ["Cámara de comercio de Duitama", "Nit: 123456789", "(Correo de la camara)", "Teléfono: (tel. camara)"],
    ]
    tabla_encabezado = Table(encabezado_data, colWidths=[180, 180, 180, 180])
    estilo_encabezado = TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    tabla_encabezado.setStyle(estilo_encabezado)
    elements.append(tabla_encabezado)

    # Datos del usuario
    usuario = request.user
    data_usuario = [["Usuario:", "Email:", "Rol:", "Cargo:"]]
    data_usuario.append([
        usuario.username,
        usuario.email,
        getattr(usuario, 'role', 'No definido'),
        getattr(usuario, 'cargo', 'No definido'),
    ])
    table_usuario = Table(data_usuario, colWidths=[180, 180, 180, 180])
    style_usuario = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table_usuario.setStyle(style_usuario)
    elements.append(table_usuario)

    # Obtener pedidos filtrados
    pedidos = get_pedidos_filtrados(request).prefetch_related('articulos__articulo')

    if not pedidos.exists():
        centered_style = ParagraphStyle(
            name="CenteredNormal",
            parent=styles["Normal"],
            alignment=TA_CENTER,
        )
        no_results = Paragraph("No se encontraron pedidos confirmados ni cancelados.", centered_style)
        elements.append(no_results)
    else:
        # Encabezado de tabla
        data_pedidos = [["ID Pedido", "Fecha", "Estado", "Registrado Por", "Artículos", "Área"]]

        # Filas de pedidos
        for pedido in pedidos:
            articulos_raw = ", ".join([
                f"{pa.cantidad} {pa.articulo.nombre}({pa.tipo})"
                for pa in pedido.articulos.all()
            ]) or 'Sin artículos'
            area_raw = pedido.articulos.first().area if pedido.articulos.exists() else 'Sin área'

            data_pedidos.append([
                wrap_text_p(str(pedido.id)),
                wrap_text_p(pedido.fecha_pedido.strftime('%d-%m-%Y')),
                wrap_text_p(pedido.get_estado_display()),
                wrap_text_p(pedido.registrado_por.username if pedido.registrado_por else 'No definido'),
                wrap_text_p(articulos_raw),
                wrap_text_p(area_raw),
            ])

        tabla_pedidos = Table(data_pedidos, colWidths=[60, 100, 100, 160, 200, 100])
        style_pedidos = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        tabla_pedidos.setStyle(style_pedidos)
        elements.append(tabla_pedidos)

    # Construir el PDF
    doc.build(elements, onFirstPage=draw_table_on_canvas, onLaterPages=draw_table_on_canvas)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Lista de pedidos Gestor CCD.pdf"'
    return response

def reporte_pedidos_excel(request):
    # Obtener todos los pedidos ordenados por fecha descendente
    pedidos = get_pedidos_filtrados(request).prefetch_related('articulos__articulo')

    # Crear libro Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos CCD"

    # Ajustar ancho de columnas
    columnas_anchos = {
        'A': 10,   # ID
        'B': 15,   # Fecha
        'C': 15,   # Estado
        'D': 25,   # Registrado por
        'E': 50,   # Artículos
        'F': 30    # Áreas
    }
    for col, width in columnas_anchos.items():
        ws.column_dimensions[col].width = width

    # Ajustar altura de título y subtítulo
    ws.row_dimensions[1].height = 50
    ws.row_dimensions[2].height = 30

    # Título principal
    ws.merge_cells('A1:F1')
    ws['A1'] = "GESTOR CCD"
    ws['A1'].font = Font(size=24, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Subtítulo
    ws.merge_cells('A2:F2')
    ws['A2'] = "Listado de Pedidos"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    # Bordes para los títulos A1:F1 y A2:F2
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=6):
        for cell in row:
            cell.border = thin_border

    # Encabezados de tabla
    headers = ['ID', 'Fecha', 'Estado', 'Registrado Por', 'Artículos', 'Áreas']
    ws.append(headers)

    # Estilo para encabezados
    header_fill = PatternFill(start_color="FF0056B3", end_color="FF0056B3", fill_type="solid")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Aplicar filtro
    ws.auto_filter.ref = "A3:F3"

    # Si no hay pedidos, mostrar mensaje
    if not pedidos.exists():
        ws.merge_cells('A4:F4')
        cell = ws['A4']
        cell.value = "No hay pedidos"
        cell.font = Font(bold=False)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    else:
        # Agregar datos de los pedidos
        for pedido in pedidos:
            articulos_raw = ", ".join([
                f"{pa.articulo.nombre} x {pa.cantidad} ({pa.tipo})"
                for pa in pedido.articulos.all()
            ]) or 'Sin artículos'

            areas_raw = ", ".join(set([
                str(pa.area) for pa in pedido.articulos.all()
            ])) or 'Sin área'

            usuario = pedido.registrado_por.username if pedido.registrado_por else 'No definido'

            ws.append([
                pedido.id,
                pedido.fecha_pedido.strftime('%Y-%m-%d'),
                pedido.get_estado_display(),
                usuario,
                articulos_raw,
                areas_raw,
            ])

        # Aplicar bordes y alineación a todas las celdas de A3:F...
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=6):
            for i, cell in enumerate(row, start=1):
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal='center',
                    vertical='center',
                    wrap_text=True if i in [4, 5, 6] else False
                )

        # Ajustar altura de filas para mejor legibilidad
        for i in range(4, ws.max_row + 1):
            ws.row_dimensions[i].height = 60

    # Generar y devolver el archivo Excel como descarga
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Reporte pedidos confirmados y cancelados.xlsx"'
    wb.save(response)
    return response

def get_pedidos_filtrados_pendientes(request):
    query = request.GET.get('q')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    pedidos = Pedido.objects.filter(estado='Pendiente').order_by('-fecha_pedido')

    if query:
        pedidos = pedidos.filter(
            Q(registrado_por__username__icontains=query) |
            Q(estado__icontains=query) |
            Q(articulos__articulo__nombre__icontains=query) |
            Q(articulos__cantidad__icontains=query) |
            Q(articulos__tipo__icontains=query) |
            Q(articulos__area__icontains=query) |
            Q(id__icontains=query)
        ).distinct()

    if fecha_inicio and fecha_fin:
        try:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d") + timedelta(days=1)
            pedidos = pedidos.filter(fecha_pedido__gte=fecha_inicio_dt, fecha_pedido__lt=fecha_fin_dt)
        except ValueError:
            pass  # opcional: puedes registrar/loggear el error si las fechas son inválidas

    return pedidos
def reporte_pedidos_pendientes_pdf(request):
    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=40,
        rightMargin=40,
        topMargin=40,
        bottomMargin=40
    )

    doc.title = "Pedidos pendientes CCD"
    doc.author = "GESTOR CCD"
    doc.subject = "Pedidos pendientes CCD"
    doc.creator = "Sistema de Gestión CCD"

    elements = []
    styles = getSampleStyleSheet()

    # Título principal
    titulo = Paragraph("Reporte de pedidos pendientes", styles["Title"])
    elements.append(titulo)

    # Encabezado institucional
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    encabezado_data = [
        ["GESTOR CCD", "Lista de usuarios", "Correo: gestorccd@gmail.com", f"Fecha: {fecha_actual}"],
        ["Cámara de comercio de Duitama", "Nit: 123456789", "(Correo de la camara)", "Teléfono: (tel. camara)"],
    ]
    tabla_encabezado = Table(encabezado_data, colWidths=[180, 180, 180, 180])
    estilo_encabezado = TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    tabla_encabezado.setStyle(estilo_encabezado)
    elements.append(tabla_encabezado)

    # Datos del usuario
    usuario = request.user
    data_usuario = [["Usuario:", "Email:", "Rol:", "Cargo:"]]
    data_usuario.append([
        usuario.username,
        usuario.email,
        getattr(usuario, 'role', 'No definido'),
        getattr(usuario, 'cargo', 'No definido'),
    ])
    table_usuario = Table(data_usuario, colWidths=[180, 180, 180, 180])
    style_usuario = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table_usuario.setStyle(style_usuario)
    elements.append(table_usuario)

    # Obtener pedidos filtrados
    pedidos = get_pedidos_filtrados_pendientes(request).prefetch_related('articulos__articulo')

    if not pedidos.exists():
        centered_style = ParagraphStyle(
            name="CenteredNormal",
            parent=styles["Normal"],
            alignment=TA_CENTER
        )
        no_results = Paragraph("No se encontraron pedidos pendientes.", centered_style)
        elements.append(Spacer(1, 20))
        elements.append(no_results)
    else:
        # Encabezado de la tabla
        data_pedidos = [["ID Pedido", "Fecha", "Estado", "Registrado Por", "Artículos", "Área"]]

        # Filas de pedidos
        for pedido in pedidos:
            articulos_raw = ", ".join([
                f"{pa.cantidad} {pa.articulo.nombre} {pa.tipo}"
                for pa in pedido.articulos.all()
            ]) or 'Sin artículos'
            articulos_text = wrap_text_p(articulos_raw)

            area_raw = pedido.articulos.first().area if pedido.articulos.exists() else 'Sin área'
            areas_text = wrap_text_p(area_raw)

            data_pedidos.append([
                wrap_text_p(str(pedido.id)),
                wrap_text_p(pedido.fecha_pedido.strftime('%d-%m-%Y')),
                wrap_text_p(pedido.get_estado_display()),
                wrap_text_p(pedido.registrado_por.username if pedido.registrado_por else 'No definido'),
                wrap_text_p(articulos_raw),
                wrap_text_p(area_raw)
            ])

        # Crear tabla de pedidos
        tabla_articulos = Table(data_pedidos, colWidths=[60, 100, 100, 160, 200, 100])
        style_articulos = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        tabla_articulos.setStyle(style_articulos)
        elements.append(tabla_articulos)

    # Construir el PDF
    doc.build(elements, onFirstPage=draw_table_on_canvas, onLaterPages=draw_table_on_canvas)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte pedidos pendientes CCD.pdf"'
    return response


def reporte_pedidos_pendientes_excel(request):
    pedidos = get_pedidos_filtrados_pendientes(request).prefetch_related('articulos__articulo')

    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos CCD"

    # Configurar anchos de columna
    columnas_anchos = {'A': 10, 'B': 15, 'C': 15, 'D': 25, 'E': 50, 'F': 30}
    for col, width in columnas_anchos.items():
        ws.column_dimensions[col].width = width

    # Alturas para título y subtítulo
    ws.row_dimensions[1].height = 60
    ws.row_dimensions[2].height = 30

    # Título principal
    ws.merge_cells('A1:F1')
    ws['A1'] = "GESTOR CCD"
    ws['A1'].font = Font(size=24, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Subtítulo con fecha
    ws.merge_cells('A2:F2')
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    ws['A2'] = f"Listado de Pedidos Pendientes - {fecha_actual}"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    # Borde del título y subtítulo
    borde = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=6):
        for cell in row:
            cell.border = borde
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Encabezados
    headers = ['ID', 'Fecha', 'Estado', 'Registrado Por', 'Artículos', 'Áreas']
    ws.append(headers)
    header_fill = PatternFill(start_color="FF0056B3", end_color="FF0056B3", fill_type="solid")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Filtro automático
    ws.auto_filter.ref = "A3:F3"

    # Agregar datos
    if not pedidos.exists():
        ws.merge_cells('A4:F4')
        cell = ws['A4']
        cell.value = "No se encontraron pedidos pendientes."
        cell.font = Font(bold=False)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = borde
    else:
        for pedido in pedidos:
            articulos_raw = ", ".join([
                f"{pa.cantidad} {pa.articulo.nombre} {pa.tipo}"
                for pa in pedido.articulos.all()
            ])
            areas_raw = ", ".join(set(str(pa.area) for pa in pedido.articulos.all()))
            usuario = pedido.registrado_por.username if pedido.registrado_por else 'No definido'
            ws.append([
                pedido.id,
                pedido.fecha_pedido.strftime('%Y-%m-%d'),
                pedido.get_estado_display() if hasattr(pedido, 'get_estado_display') else pedido.estado,
                usuario,
                articulos_raw or 'Sin artículos',
                areas_raw or 'Sin área',
            ])

    # Estilos de celdas: bordes y alineación
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=6):
        for i, cell in enumerate(row, start=1):
            cell.border = borde
            cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True if i in [4, 5, 6] else False
            )

    # Altura dinámica para filas de contenido
    for i in range(4, ws.max_row + 1):
        ws.row_dimensions[i].height = 60

    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Reporte pedidos pendientes papeleria.xlsx"'
    wb.save(response)
    return response
#PDF Y XSLS DE BAJO STOCK
def obtener_articulos_bajo_stock(request):
    query = request.GET.get('q')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    articulos = Articulo.objects.filter(cantidad__lt=10)

    if query:
        articulos = articulos.filter(
            Q(id__icontains=query) |
            Q(nombre__icontains=query) |
            Q(marca__icontains=query) |
            Q(observacion__icontains=query) |
            Q(tipo__icontains=query) |
            Q(precio__icontains=query) |
            Q(proveedor__icontains=query) |
            Q(cantidad__icontains=query) |
            Q(registrado_por__username__icontains=query)
        )

    if fecha_inicio and fecha_fin:
        articulos = articulos.filter(fecha_registro__range=[fecha_inicio, fecha_fin])

    return articulos
def reporte_articulo_bajo_stock_pdf(request):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                            leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)

    doc.title = "Listado de BAJO STOCK CCD"
    doc.author = "CCD"
    doc.subject = "Listado de bajo stock"
    doc.creator = "Sistema de Gestión CCD"

    elements = []
    styles = getSampleStyleSheet()

    # Título
    titulo = Paragraph("REPORTE DE BAJO STOCK", styles["Title"])
    elements.append(titulo)

    # Encabezado empresa
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    encabezado_data = [
        ["GESTOR CCD", "Lista de artículos", "Correo:", f"Fecha: {fecha_actual}"],
        ["Cámara de comercio de Duitama", "Nit: 123456789", "contacto@gestorccd.com", "Teléfono: (123) 456-7890"],
    ]
    tabla_encabezado = Table(encabezado_data, colWidths=[180, 180, 180, 180])
    estilo_encabezado = TableStyle([
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    tabla_encabezado.setStyle(estilo_encabezado)
    elements.append(tabla_encabezado)

    # Tabla usuario
    usuario = request.user
    data_usuario = [["Usuario", "Email", "Rol", "Cargo"]]
    data_usuario.append([
        usuario.username,
        usuario.email,
        getattr(usuario, 'role', 'No definido'),
        getattr(usuario, 'cargo', 'No definido'),
    ])
    table_usuario = Table(data_usuario, colWidths=[180, 180, 180, 180])
    style_usuario = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table_usuario.setStyle(style_usuario)
    elements.append(table_usuario)

    # Obtener artículos filtrados
    articulos_filtrados = obtener_articulos_bajo_stock(request)

    if not articulos_filtrados.exists():
        centered_style = ParagraphStyle(
            name="CenteredNormal",
            parent=styles["Normal"],
            alignment=TA_CENTER
        )
        no_results = Paragraph("No se encontraron artículos con bajo stock.", centered_style)
        elements.append(Spacer(1, 20))
        elements.append(no_results)
    else:
        # Tabla artículos
        data_articulos = [["ID", "Nombre", "Marca", "Tipo", "Precio", "Cantidad", "Observación"]]
        
        for articulo in articulos_filtrados:
            data_articulos.append([
                wrap_text(str(articulo.id)),
                wrap_text(articulo.nombre),
                wrap_text(articulo.marca),
                wrap_text(articulo.tipo),
                wrap_text("{:,}".format(articulo.precio)),
                wrap_text(str(articulo.cantidad)),
                wrap_text(articulo.observacion),
            ])

        tabla_articulos = Table(data_articulos, colWidths=[20, 140, 100, 130, 80, 70, 180])
        style_articulos = TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5564eb")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        tabla_articulos.setStyle(style_articulos)
        elements.append(tabla_articulos)

    # Marca de agua
    doc.build(elements, onFirstPage=draw_table_on_canvas, onLaterPages=draw_table_on_canvas)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Listado bajo stock CCD.pdf"'
    return response
def reporte_articulo_bajo_stock_excel(request):
    # Obtener artículos filtrados usando la función reutilizable
    articulos = obtener_articulos_bajo_stock(request)

    # Crear archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Listado de Artículos CCD"

    # Configurar columnas
    column_widths = [10, 40, 20, 30, 20, 15, 33]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    # Altura de filas
    ws.row_dimensions[1].height = 60
    ws.row_dimensions[2].height = 30

    # Título principal
    ws.merge_cells('A1:G1')
    ws['A1'] = "GESTOR CCD"
    ws['A1'].font = Font(size=24, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Subtítulo con fecha actual
    ws.merge_cells('A2:G2')
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    ws['A2'] = f"Listado de Artículos - {fecha_actual}"
    ws['A2'].font = Font(size=18)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    # Encabezados
    headers = ["ID", "Nombre", "Marca", "Tipo", "Precio", "Cantidad", "Observación"]
    ws.append(headers)

    header_fill = PatternFill(start_color="FF0056B3", end_color="FF0056B3", fill_type="solid")
    for col in range(1, 8):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Bordes
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Agregar datos o mensaje
    if not articulos.exists():
        ws.merge_cells('A4:G4')
        cell = ws['A4']
        cell.value = "No se encontraron artículos en bajo stock."
        cell.font = Font(bold=False)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    else:
        for articulo in articulos:
            precio_formateado = f"${articulo.precio:,.2f}" if articulo.precio else "-"
            ws.append([
                articulo.id,
                str(articulo.nombre),
                str(articulo.marca),
                str(articulo.tipo),
                precio_formateado,
                articulo.cantidad,
                str(articulo.observacion or ""),
            ])

    # Estilo a título y subtítulo
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=7):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Estilo a encabezados y datos
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=7):
        for i, cell in enumerate(row, start=1):
            cell.border = border
            if i < 7:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Filtro automático
    ws.auto_filter.ref = "A3:G3"

    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Listado_articulos_bajo_stock.xlsx"'
    wb.save(response)
    return response